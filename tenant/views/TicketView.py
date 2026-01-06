import logging
import os
import re
from datetime import datetime, timedelta
from io import BytesIO

import uuid
from openpyxl import Workbook
from django.http import HttpResponse

from django.db.utils import IntegrityError
from django.db import transaction
from django.forms import ValidationError
from django.http import Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from django.db.models import Q, Count
from RNSafarideskBack import settings
from RNSafarideskBack.settings import FILE_BASE_URL, FILE_URL
from tenant.models import Ticket, TicketCategories, Department, TicketAttachment, Task, TicketWatchers
from tenant.models.SlaXModel import SLA, SLATarget, SLAConfiguration # Import SLA and SLATarget from SlaXModel
from tenant.models.TicketModel import TicketComment, TicketReplayAttachment, CommentLike, CommentReply, \
    CommentReplyLike, TicketReopen
from tenant.models.ConfigModel import TicketConfig
from tenant.serializers.TicketSerializer import TicketAddComment, TicketAssign, TicketCategorySerializer, \
    TicketSerializer, TicketUpdateStatus, TicketsList, TicketAddWatchers, TicketAddTags, TicketCategoryUpdateSerializer, \
    TicketMergeRequestSerializer
from tenant.serializers.EmailMessageRecordSerializer import EmailMessageRecordSerializer
from users.models import Users
from util.Constants import PRIORITY_DURATION
from util.Helper import Helper
from shared.tasks import create_notification_task, send_mention_email_notification
from shared.workers.Ticket import ticket_claim_task, _send_ticket_notification
from shared.services.notification_preferences import NotificationSettingsService
from tenant.services.contact_linker import link_or_create_contact
logger = logging.getLogger(__name__)


def _sanitize_text_for_mysql(text: str) -> str:
    """
    Remove 4-byte Unicode characters (like emojis) when the database engine
    is MySQL and the connection isn't configured to use utf8mb4. This prevents
    MySQL errors like "Incorrect string value: '\xF0\x9F...'" on insert.

    We only sanitize when a non-utf8mb4 MySQL is used; otherwise, return text unchanged.
    """
    try:
        engine = settings.DATABASES.get('default', {}).get('ENGINE', '')
    except Exception:
        engine = ''

    # Only apply for MySQL with incompatible charset
    if 'mysql' in engine.lower():
        opts = settings.DATABASES.get('default', {}).get('OPTIONS', {}) or {}
        charset = opts.get('charset', '')
        if charset and charset.lower() == 'utf8mb4':
            return text

        # Remove non-BMP characters (codepoints > 0xFFFF) - emojis are outside BMP
        if text:
            try:
                # Keep BMP (<= 0xFFFF) characters only
                return ''.join(ch for ch in text if ord(ch) <= 0xFFFF)
            except Exception:
                # Fallback: replace unknown chars to avoid crash
                return text.encode('utf-8', errors='replace').decode('utf-8')

    return text

helper = Helper()

def _build_storage_paths(business, subfolder="files"):
    """
    Return (dir_path, url_base) for storing files under MEDIA_ROOT/files/<business_id>/...
    """
    base_media = settings.MEDIA_ROOT
    if business:
        dir_path = os.path.join(base_media, subfolder, str(business.id))
        url_base = f"{FILE_URL}/{business.id}" if FILE_URL else f"{getattr(settings, 'FILE_BASE_URL', '')}/{business.id}"
    else:
        dir_path = os.path.join(base_media, subfolder)
        url_base = FILE_URL or getattr(settings, 'FILE_BASE_URL', '')
    os.makedirs(dir_path, exist_ok=True)
    return dir_path, url_base


def _extract_and_save_base64_images(html_content: str, comment, business) -> str:
    """
    Extract base64-encoded images from HTML content, save them as files,
    create TicketReplayAttachment records, and replace base64 src with file URLs.
    
    Args:
        html_content: HTML string potentially containing base64 images
        comment: TicketComment instance to attach files to
        business: Business instance for storage path
    
    Returns:
        Modified HTML with base64 images replaced by file URLs
    """
    import base64
    
    if not html_content:
        return html_content
    
    # Regex to find base64 images in img src attributes
    # Matches: src="data:image/png;base64,..." or src='data:image/jpeg;base64,...'
    base64_pattern = re.compile(
        r'src=["\']data:image/(png|jpg|jpeg|gif|webp);base64,([A-Za-z0-9+/=]+)["\']',
        re.IGNORECASE
    )
    
    modified_html = html_content
    matches = list(base64_pattern.finditer(html_content))
    
    for match in matches:
        try:
            image_type = match.group(1).lower()
            base64_data = match.group(2)
            original_src = match.group(0)
            
            # Decode base64 to binary
            image_data = base64.b64decode(base64_data)
            
            # Generate unique filename
            file_extension = f".{image_type}" if image_type != 'jpg' else '.jpeg'
            unique_filename = f"{uuid.uuid4()}{file_extension}"
            
            # Build storage paths
            properties_dir, url_base = _build_storage_paths(business, subfolder='files')
            file_path = os.path.join(properties_dir, unique_filename)
            
            # Save file
            with open(file_path, 'wb') as f:
                f.write(image_data)
            
            # Generate URL
            file_url_base = url_base or FILE_URL or getattr(settings, 'FILE_BASE_URL', '')
            if file_url_base:
                file_url = f"{file_url_base.rstrip('/')}/{unique_filename}"
            else:
                file_url = unique_filename
            
            # Create attachment record
            TicketReplayAttachment.objects.create(
                comment=comment,
                file_url=file_url,
                filename=f"pasted_image_{uuid.uuid4().hex[:8]}{file_extension}"
            )
            
            # Replace base64 src with file URL in HTML
            new_src = f'src="{file_url}"'
            modified_html = modified_html.replace(original_src, new_src)
            
            logger.info(f"Extracted and saved base64 image: {unique_filename}")
            
        except Exception as e:
            logger.error(f"Failed to extract base64 image: {e}")
            # Continue with other images, don't fail the whole operation
            continue
    
    return modified_html


class TicketPagination(PageNumberPagination):
    page_size = 12  # default items per page
    page_size_query_param = 'page_size'  # allow overriding via ?page_size=XX (optional)
    max_page_size = 100

class TicketCategoryView(viewsets.ModelViewSet):
    queryset = TicketCategories.objects.all()
    serializer_class = TicketCategorySerializer
    permission_classes = [IsAuthenticated]
    pagination_class = PageNumberPagination

    def get_serializer_class(self):
        if self.action in ['update', 'partial_update']:
            return TicketCategoryUpdateSerializer
        return TicketCategorySerializer

    def create(self, request, *args, **kwargs):
        """Create a new ticket category"""
        name = request.data.get("name")
        description = request.data.get("description")
        department_id = request.data.get("department")
        
        # Check for duplicate name within department (or business if no department)
        filter_kwargs = {'name': name}
        if department_id:
            filter_kwargs['department_id'] = department_id
        
        if TicketCategories.objects.filter(**filter_kwargs).exists():
            return Response({
                "message": "Ticket category exists",
            }, status=401)

        cat = TicketCategories.objects.create(
            name=name,
            description=description,
            department_id=department_id if department_id else None
        )

        return Response({
            "message": "Ticket category created successfully",
        }, status=status.HTTP_201_CREATED)

    def list(self, request, *args, **kwargs):
        """List all ticket categories, optionally filtered by department"""
        queryset = TicketCategories.objects.for_business()
        
        # Filter by department if provided
        department_id = request.query_params.get('department')
        if department_id:
            queryset = queryset.filter(department_id=department_id)
        
        pagination = request.query_params.get('pagination', 'yes').lower()

        if pagination != 'no':
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        # Unpaginated response
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


    def retrieve(self, request, pk=None, *args, **kwargs):
        """Get a specific ticket category by ID"""
        instance = get_object_or_404(self.get_queryset(), pk=pk)
        serializer = self.get_serializer(instance)
        return Response({
            "message": "Ticket category retrieved successfully",
            "data": serializer.data
        }, status=status.HTTP_200_OK)

    def update(self, request, pk=None, *args, **kwargs):
        """Update a ticket category"""
        instance = get_object_or_404(self.get_queryset(), pk=pk)
        serializer = self.get_serializer(instance, data=request.data, partial=True)

        if serializer.is_valid():
            serializer.save()
            return Response({
                "message": "Ticket category updated successfully",
                "data": serializer.data
            }, status=status.HTTP_200_OK)

        return Response({
            "message": "Validation failed",
            "errors": serializer.errors
        }, status=status.HTTP_400_BAD_REQUEST)


    def destroy(self, request, pk=None, *args, **kwargs):
        """Deactivate/activate a ticket category"""
        instance = get_object_or_404(self.get_queryset(), pk=pk)
        instance.is_active = not instance.is_active
        instance.save()
        status_message = "activated" if instance.is_active else "deactivated"
        return Response({
            "message": f"Ticket category {status_message} successfully"
        }, status=status.HTTP_200_OK)

class TicketView(viewsets.ModelViewSet):
    queryset = Ticket.objects.all()
    pagination_class = TicketPagination

    def get_permissions(self):
        if self.action in ["loadActivityStream", "loadAttachments", "add_comment"]:
            return [AllowAny()]
        return [IsAuthenticated()]

    def get_serializer_class(self):
        if self.action == 'assign':
            return TicketAssign
        if self.action == 'merge':
            return TicketMergeRequestSerializer
        
        if self.action in ['list', 'read_by_ticket_id', 'my_tickets', 'my_customer_tickets', 'profile_ticket_counts']:
            return TicketsList
        
        if self.action == 'update_status':
            return TicketUpdateStatus
        if self.action == 'add_comment':
            return TicketAddComment
        if self.action == 'add_ticket_watchers':
            return TicketAddWatchers
        if self.action == 'add_ticket_tags':
            return TicketAddTags
        return TicketSerializer

    def loadActivityStream(self, request, *args, **kwargs):
        id = kwargs.get('id')

        try:
            # Get ticket with related data using select_related and prefetch_related for optimization
            ticket = Ticket.objects.select_related(
                'category',
                'department',
                'assigned_to'
            ).prefetch_related(
                'comments__author',
                'comments__attachment',
                'comments__replies__author',
                'attachments',
                'activities__user'
            ).get(
                id=id
            )
            merged_children = ticket.merged_children.prefetch_related(
                'comments__author',
                'comments__attachment',
                'comments__replies__author',
                'attachments',
                'activities__user'
            ).all()
        except Ticket.DoesNotExist:
            return Response({
                "error": "Ticket not found"
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Collect all activities (comments + system events)
        all_activities = []

        def _append_comments(src_ticket):
            for comment in src_ticket.comments.all():
                attachments = []
                for a in comment.attachment.all():
                    attachments.append({
                        "id": a.id,
                        "file_url": a.file_url,
                        "filename": a.filename,  # Include original filename
                    })

                replies_data = []
                for reply in comment.replies.all():
                    replies_data.append({
                        "id": reply.id,
                        "content": reply.content,
                        "created_at": reply.created_at,
                        "updated_at": reply.updated_at,
                        "is_internal": reply.is_internal,
                        "likes_count": reply.likes_count,
                        "author": {
                            "id": reply.author.id,
                            "avatar": reply.author.avatar_url if reply.author else "",
                            "name": reply.author.full_name() if reply.author else 'Anonymous User',
                            "email": reply.author.email if reply.author else 'Anonymous User',
                        }
                    })

                comment_data = {
                    "id": comment.id,
                    "type": "comment",
                    "content": comment.content,
                    "created_at": comment.created_at,
                    "updated_at": comment.updated_at,
                    "is_internal": comment.is_internal,
                    "is_solution": comment.is_solution,
                    "attachments": attachments if attachments else None,
                    "author": {
                        "id": comment.author.id,
                        "avatar": comment.author.avatar_url if comment.author else "",
                        "name": comment.author.full_name() if comment.author else 'Anonymous User',
                        "email": comment.author.email if comment.author else 'Anonymous User',
                    } if comment.author else {
                        "id": 0,
                        "avatar": "",
                        "name": src_ticket.creator_name,
                        "email": src_ticket.creator_email,
                    },
                    "replies": replies_data,
                    "_sort_timestamp": comment.created_at,
                    "source_ticket": {
                        "id": src_ticket.id,
                        "ticket_id": src_ticket.ticket_id,
                        "title": src_ticket.title,
                    },
                    # Email recipient fields (for email replies)
                    "email_to": comment.email_to,
                    "email_cc": comment.email_cc,
                    "email_bcc": comment.email_bcc,
                }
                all_activities.append(comment_data)

        def _append_activities(src_ticket):
            for activity in src_ticket.activities.all():
                activity_data = {
                    "id": activity.id,
                    "type": "system_event",
                    "activity_type": activity.activity_type,
                    "user": {
                        "id": activity.user.id,
                        "avatar": activity.user.avatar_url if activity.user else "",
                        "name": activity.user.full_name() if activity.user else 'System',
                        "email": activity.user.email if activity.user else '',
                    },
                    "timestamp": activity.timestamp,
                    "description": activity.description,
                    "old_value": activity.old_value,
                    "new_value": activity.new_value,
                    "_sort_timestamp": activity.timestamp,
                    "source_ticket": {
                        "id": src_ticket.id,
                        "ticket_id": src_ticket.ticket_id,
                        "title": src_ticket.title,
                    }
                }
                all_activities.append(activity_data)

        _append_comments(ticket)
        _append_activities(ticket)
        # Do not append merged child activity streams to avoid duplication/confusion.
        # Users can follow hyperlinks to merged tickets for their history.
        # for child in merged_children:
        #     _append_comments(child)
        #     _append_activities(child)
        
        # Sort by timestamp (chronological order)
        all_activities.sort(key=lambda x: x['_sort_timestamp'])
        
        # Remove the _sort_timestamp helper field
        for activity in all_activities:
            del activity['_sort_timestamp']

        return Response(all_activities)

    def loadAttachments(self, request, *args, **kwargs):
        id = kwargs.get('id')

        try:
            ticket = Ticket.objects.select_related(
                'category',
                'department',
                'assigned_to'
            ).prefetch_related(
                'comments__author',
                'comments__attachment',
                'attachments',
                'activities__user'
            ).get(
                id=id
            )
            merged_children = ticket.merged_children.prefetch_related(
                'comments__author',
                'comments__attachment',
                'attachments'
            ).all()
        except Ticket.DoesNotExist:
            return Response({
                "error": "Ticket not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Ticket-level attachments
        attachment_urls = [
            {
                "id": attachment.id,
                "file_url": attachment.file_url,
                "ticket_id": ticket.ticket_id,
            }
            for attachment in ticket.attachments.all()
        ]

        for comment in ticket.comments.all():
            comment_attachment_urls = [
                {
                    "id": a.id,
                    "file_url": a.file_url,
                    "ticket_id": ticket.ticket_id,
                }
                for a in comment.attachment.all()
            ]
            attachment_urls.extend(comment_attachment_urls)

        for child in merged_children:
            attachment_urls.extend(
                [
                    {
                        "id": a.id,
                        "file_url": a.file_url,
                        "ticket_id": child.ticket_id,
                    }
                    for a in child.attachments.all()
                ]
            )
            for comment in child.comments.all():
                attachment_urls.extend(
                    [
                        {
                            "id": a.id,
                            "file_url": a.file_url,
                            "ticket_id": child.ticket_id,
                        }
                        for a in comment.attachment.all()
                    ]
                )

        # Remove duplicates (optional)
        unique_urls = []
        seen = set()
        for item in attachment_urls:
            key = (item.get("id"), item.get("file_url"), item.get("ticket_id"))
            if key in seen:
                continue
            seen.add(key)
            unique_urls.append(item)

        return Response(unique_urls, status=status.HTTP_200_OK)

    def loadTasks(self, request, *args, **kwargs):
        id = kwargs.get('id')

        try:
            ticket = Ticket.objects.select_related(
                'category',
                'department',
                'assigned_to'
            ).prefetch_related(
                'comments__author',
                'comments__attachment',
                'attachments',
                'activities__user'
            ).get(
                id=id
            )
        except Ticket.DoesNotExist:
            return Response({
                "error": "Ticket not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Get tasks linked to this ticket
        tasks = Task.objects.prefetch_related(
            'attachments',
            'comments__attachment',
            'comments__author'
        ).filter(
            linked_ticket=ticket,
            is_archived=False,
            is_deleted=False,
        )

        tasks_data = []

        for task in tasks:
            # Task-level attachments
            task_attachments = [att.file_url for att in task.attachments.all()]

            # Task comments with their attachments
            comments_data = []
            for comment in task.comments.all():
                comment_attachments = [a.file_url for a in comment.attachment.all()]
                comments_data.append({
                    "id": comment.id,
                    "content": comment.content,
                    "author": comment.author.full_name() if comment.author else "Anonymous",
                    "created_at": comment.created_at,
                    "attachments": comment_attachments
                })

            tasks_data.append({
                "id": task.id,
                "task_trackid": task.task_trackid,
                "title": task.title,
                "description": task.description,
                "priority": task.priority,
                "status": task.task_status,
                "assigned_to": task.assigned_to.full_name() if task.assigned_to else None,
                "due_date": task.due_date,
                "completed_at": task.completed_at,
                "attachments": task_attachments,
                "comments": comments_data
            })

        return Response(tasks_data)

    def getSla(self, request, *args, **kwargs):
        """
        Get comprehensive SLA details for a specific ticket.
        """
        try:
            # Check if SLA is allowed
            sla_config = SLAConfiguration.objects.filter(pk=1).first()
            if not sla_config or not sla_config.allow_sla:
                return Response({
                    "message": "SLA tracking is currently disabled",
                    "sla_enabled": False
                }, status=status.HTTP_200_OK)
            
            ticket = Ticket.objects.for_business().get(id=kwargs.get("id"))
            sla_analysis_data = ticket.sla_analysis()
            sla_status_data = ticket.get_sla_status()

            # Calculate percentages for business hours elapsed and system hours elapsed
            business_hours_elapsed = sla_analysis_data.get('business_hours_elapsed', 0)
            system_hours_elapsed = sla_analysis_data.get('system_hours_elapsed', 0)
            
            total_business_hours = sla_analysis_data.get('total_business_hours_for_resolution')
            total_system_hours = sla_analysis_data.get('total_system_hours_for_resolution')

            if total_business_hours is not None and total_business_hours > 0:
                sla_analysis_data['business_hours_elapsed_percentage'] = (business_hours_elapsed / total_business_hours) * 100
            else:
                sla_analysis_data['business_hours_elapsed_percentage'] = 0.0 # Ensure float for consistency

            if total_system_hours is not None and total_system_hours > 0:
                sla_analysis_data['system_hours_elapsed_percentage'] = (system_hours_elapsed / total_system_hours) * 100
            else:
                sla_analysis_data['system_hours_elapsed_percentage'] = 0.0 # Ensure float for consistency

            response_data = {
                "ticket_id": ticket.ticket_id,
                "sla_analysis": sla_analysis_data,
                "sla_status": sla_status_data,
                "is_sla_breached": ticket.is_sla_breached,
                "is_sla_paused": ticket.is_sla_paused,
            }
            return Response(response_data, status=status.HTTP_200_OK)
        except Ticket.DoesNotExist:
            return Response({
                "message": "Ticket not found."
            }, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.error(f"Error retrieving SLA details for ticket ID {kwargs.get('id')}: {str(e)}")
            return Response({
                "message": "An error occurred while retrieving SLA details.",
                "details": str(e) if settings.DEBUG else "Please contact support"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def read_by_ticket_id(self, request, *args, **kwargs):
        ticket_id = kwargs.get('ticket_id')
        try:
            ticket = Ticket.objects.select_related(
                'category', 
                'department', 
                'assigned_to'
            ).prefetch_related(
                'comments__author',
                'comments__attachment',
                'attachments',
                'activities__user',
                'email_messages',
            ).get(
                ticket_id=ticket_id
            )
            
            # Apply department-based visibility restrictions for non-admin agents
            user = request.user
            if user.role and user.role.name == 'agent':
                # Check if the agent belongs to the ticket's department
                agent_departments = user.department.all()
                if ticket.department not in agent_departments:
                    return Response({
                        "error": "You do not have permission to view this ticket"
                    }, status=status.HTTP_403_FORBIDDEN)
                    
        except Ticket.DoesNotExist:
            return Response({
                "error": "Ticket not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Mark ticket as opened and clear new reply badge
        update_fields = []
        if not ticket.is_opened:
            ticket.is_opened = True
            update_fields.append('is_opened')
        if ticket.has_new_reply:
            ticket.has_new_reply = False
            update_fields.append('has_new_reply')
        if update_fields:
            ticket.save(update_fields=update_fields)


        # Calculate SLA status based on first response and resolution dates
        sla_status = ticket.get_sla_status()
        breached = False

        if sla_status and sla_status.get('has_sla'):
            current_time = timezone.now()

            # If ticket is resolved, SLA is not breached regardless of first response status
            if ticket.status == 'closed':
                breached = False
            else:
                # Check first response status
                first_response_info = sla_status.get('first_response', {})
                if first_response_info.get('due_time'):
                    if ticket.first_response_at:
                        # First response has been made - check resolution status
                        resolution_info = sla_status.get('resolution', {})
                        if (resolution_info.get('due_time') and
                            current_time > resolution_info['due_time']):
                            breached = True
                    elif current_time > first_response_info['due_time']:
                        # First response is overdue and not yet responded to
                        breached = True

        ticket_data = {
            "id": ticket.id,
            "ticket_id": ticket.ticket_id,
            "title": ticket.title,
            "description": ticket.description,
            "status": ticket.status,
            "priority": ticket.priority,
            "created_at": ticket.created_at,
            "updated_at": ticket.updated_at,
            "due_date": ticket.calculate_sla_due_times()['resolution_due'] if ticket.calculate_sla_due_times() else None,
            "resolved_at": ticket.resolved_at,
            "is_public": ticket.is_public,
            "tags": ticket.get_tags_list(),
            "breached": breached,
            # "is_overdue": helper.format_datetime(ticket.is_overdue),


            #
            # Creator information
            "creator_name": ticket.creator_name,
            "creator_email": ticket.creator_email,
            "creator_phone": ticket.creator_phone,

            # Email messages
            "email_messages": EmailMessageRecordSerializer(ticket.email_messages.all(), many=True).data,

            # Related objects
            "category": {
                "id": ticket.category.id,
                "name": ticket.category.name,
                "description": ticket.category.description
            } if ticket.category else None,

            "department": {
                "id": ticket.department.id,
                "name": getattr(ticket.department, 'name', 'N/A')
            } if ticket.department else None,

            "assigned_to": {
                "id": ticket.assigned_to.id,
                "name": ticket.assigned_to.get_full_name(),
                "email": ticket.assigned_to.email,
                "avatar": ticket.assigned_to.avatar_url,
            } if ticket.assigned_to else None,
            "is_merged": ticket.is_merged,
            "merged_into": {
                "id": ticket.merged_into.id,
                "ticket_id": ticket.merged_into.ticket_id,
                "title": ticket.merged_into.title,
            } if ticket.merged_into else None,
            "merged_children": [
                {
                    "id": child.id,
                    "ticket_id": child.ticket_id,
                    "title": child.title,
                } for child in ticket.merged_children.all()
            ] if hasattr(ticket, "merged_children") else [],
        }

        # Add linked tasks count
        from tenant.models import Task
        ticket_data['linked_tasks_count'] = Task.objects.filter(
            linked_ticket=ticket, 
            status='A'
        ).count()

        # Add unread activity count for current user
        from tenant.models import ActivityReadStatus
        activity_ids = ticket.activities.values_list('id', flat=True)
        read_activity_ids = ActivityReadStatus.objects.filter(
            activity_id__in=activity_ids,
            user=request.user
        ).values_list('activity_id', flat=True)
        ticket_data['unread_activity_count'] = len(activity_ids) - len(read_activity_ids)

        
        # Compile final response
        response_data = {
            "ticket": ticket_data,
        }
        
        return Response(response_data, status=status.HTTP_200_OK)

    def list(self, request, *args, **kwargs):
        """List all tickets, optionally without pagination and with view-based filtering"""
        view = request.query_params.get('view', None)
        queryset = Ticket.objects.for_business()
        
        # Apply department-based visibility restrictions for non-admin agents
        user = request.user
        if user.role and user.role.name == 'agent':
            # Get all departments the agent belongs to
            agent_departments = user.department.all()
            # Filter tickets to only show those in the agent's departments
            queryset = queryset.filter(department__in=agent_departments)

        # Exclude archived, deleted, and merged tickets from normal views
        if view not in ['archived', 'trash', 'merged']:
            queryset = queryset.filter(is_archived=False, is_deleted=False, is_merged=False)
        
        if view:
            # View names now match the count keys in ticket_counts endpoint
            if view == 'all_unresolved':
                queryset = queryset.exclude(status__in=['closed']).filter(is_merged=False)
            elif view == 'all_tickets':
                queryset = queryset.filter(is_merged=False)
            elif view == 'all_unassigned':
                queryset = queryset.filter(assigned_to__isnull=True, is_merged=False)
            elif view == 'all_resolved':
                queryset = queryset.filter(status__in=['closed'], is_merged=False)
            elif view == 'my_overdue':
                queryset = queryset.filter(assigned_to=request.user, due_date__lt=timezone.now(), is_merged=False).exclude(status__in=['closed'])
            elif view == 'my_unresolved':
                queryset = queryset.filter(assigned_to=request.user, is_merged=False).exclude(status__in=['closed'])
            elif view == 'my_resolved':
                queryset = queryset.filter(assigned_to=request.user, status__in=['closed'], is_merged=False)
            elif view == 'requested_by_me':
                queryset = queryset.filter(created_by=request.user, is_merged=False)
            elif view == 'sla_breached':
                # This is inefficient. A better approach is needed for production.
                breached_tickets = [ticket.id for ticket in queryset.filter(is_merged=False) if ticket.is_sla_breached]
                queryset = queryset.filter(id__in=breached_tickets)
            elif view == 'reopened':
                # Tickets that have been reopened (have at least one TicketReopen record)
                reopened_ticket_ids = TicketReopen.objects.values_list('ticket_id', flat=True).distinct()
                queryset = queryset.filter(id__in=reopened_ticket_ids, is_merged=False)
            elif view == 'archived':
                # Archived tickets: is_archived=True AND is_deleted=False
                queryset = queryset.filter(is_archived=True, is_deleted=False)
            elif view == 'trash':
                # Trash/Deleted tickets: is_deleted=True (regardless of is_archived)
                queryset = queryset.filter(is_deleted=True)
            elif view == 'merged':
                queryset = queryset.filter(is_merged=True)

        # Apply search filter
        search_query = request.query_params.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(ticket_id__icontains=search_query) |
                Q(title__icontains=search_query) |
                Q(description__icontains=search_query) |
                Q(creator_name__icontains=search_query) |
                Q(creator_email__icontains=search_query)
            )

        # Apply individual field filters
        status_filter = request.query_params.get('status')
        if status_filter and status_filter != 'all':
            queryset = queryset.filter(status=status_filter)
        
        priority_filter = request.query_params.get('priority')
        if priority_filter and priority_filter != 'all':
            queryset = queryset.filter(priority=priority_filter)
        
        assigned_to_filter = request.query_params.get('assigned_to')
        if assigned_to_filter:
            if assigned_to_filter == 'unassigned':
                queryset = queryset.filter(assigned_to__isnull=True)
            elif assigned_to_filter != 'all':
                queryset = queryset.filter(assigned_to_id=assigned_to_filter)
        
        department_filter = request.query_params.get('department')
        if department_filter and department_filter != 'all':
            queryset = queryset.filter(department_id=department_filter)
        
        category_filter = request.query_params.get('category')
        if category_filter and category_filter != 'all':
            queryset = queryset.filter(category_id=category_filter)
        
        # Date range filters
        date_from = request.query_params.get('date_from')
        if date_from:
            try:
                from_date = timezone.datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__gte=from_date)
            except (ValueError, AttributeError):
                pass
        
        date_to = request.query_params.get('date_to')
        if date_to:
            try:
                to_date = timezone.datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__lte=to_date)
            except (ValueError, AttributeError):
                pass

        pagination = request.query_params.get('pagination', 'yes').lower()

        if pagination != 'no':
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        # Unpaginated fallback
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def ticket_counts(self, request, *args, **kwargs):
        """Return counts of tickets for the sidebar views
        
        Counts exclude archived and deleted tickets from active counts.
        Archived and deleted are tracked separately and are mutually exclusive.
        """
        
        base_queryset = Ticket.objects.for_business()
        
        # Apply department-based visibility restrictions for non-admin agents
        user = request.user
        if user.role and user.role.name == 'agent':
            # Get all departments the agent belongs to
            agent_departments = user.department.all()
            # Filter tickets to only show those in the agent's departments
            base_queryset = base_queryset.filter(department__in=agent_departments)
        
        # Exclude archived, deleted, and merged from active counts to prevent double counting
        active_queryset = base_queryset.filter(is_archived=False, is_deleted=False, is_merged=False)
        my_active_queryset = active_queryset.filter(assigned_to=request.user)
        
        # For archived tickets (exclude deleted)
        archived_queryset = base_queryset.filter(is_archived=True, is_deleted=False)
        
        # For trash/deleted tickets
        trash_queryset = base_queryset.filter(is_deleted=True)

        # This is inefficient for SLA breached. A better approach is needed for production.
        sla_breached_count = sum(1 for ticket in active_queryset if ticket.is_sla_breached)

        # Count tickets that have been reopened (tickets with at least one reopen record)
        reopened_ticket_ids = TicketReopen.objects.values_list('ticket_id', flat=True).distinct()
        reopened_count = active_queryset.filter(id__in=reopened_ticket_ids).count()

        counts = {
            # Active tickets (excluding archived and deleted)
            # Keys match frontend view names in viewConfigs.ts
            'all_tickets': active_queryset.count(),
            'all_unassigned': active_queryset.filter(assigned_to__isnull=True).count(),
            'all_unresolved': active_queryset.exclude(status__in=['closed']).count(),
            'all_resolved': active_queryset.filter(status__in=['closed']).count(),
            'reopened': reopened_count,
            'merged': base_queryset.filter(is_merged=True).count(),
            
            # My tickets (excluding archived and deleted)
            'my_overdue': my_active_queryset.filter(due_date__lt=timezone.now()).exclude(status__in=['closed']).count(),
            'my_unresolved': my_active_queryset.exclude(status__in=['closed']).count(),
            'my_resolved': my_active_queryset.filter(status__in=['closed']).count(),
            
            # Special status tickets
            'requested_by_me': active_queryset.filter(created_by=request.user).count(),
            'sla_breached': sla_breached_count,
            
            # Archived and Trash (mutually exclusive)
            'archived': archived_queryset.count(),
            'trash': trash_queryset.count(),
            
            # Suspended emails and spam (if supported by model)
            'suspended_emails': 0,  # TODO: Add suspended_emails field to Ticket model if needed
            'spam': 0,  # TODO: Add spam_marked field to Ticket model if needed
        }
        
        return Response(counts, status=status.HTTP_200_OK)

    def my_tickets(self, request, *args, **kwargs):
        """List all tickets for user (by departments or as watcher), optionally without pagination"""
        user_departments = request.user.department.all()

        queryset = Ticket.objects.filter(
            Q(department__in=user_departments) |
            Q(watchers__watcher=request.user)
        ).filter(is_merged=False).order_by('-id').distinct()

        pagination = request.query_params.get('pagination', 'yes').lower()

        if pagination != 'no':
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        # Unpaginated fallback
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def my_customer_tickets(self, request, *args, **kwargs):
        """List all tickets for the authenticated user, with optional pagination and search"""
        queryset = Ticket.objects.filter(
            created_by=request.user
        ).filter(is_merged=False).order_by('-id')

        # Handle search
        search = request.query_params.get('search')
        if search:
            queryset = queryset.filter(
                Q(ticket_id__icontains=search) |
                Q(title__icontains=search) |
                Q(description__icontains=search)
            )

        # Handle pagination
        pagination = request.query_params.get('pagination', 'yes').lower()
        if pagination != 'no':
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        # Unpaginated fallback
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def profile_ticket_counts(self, request, *args, **kwargs):
        """Return counts of tickets and tasks assigned to or completed by the authenticated user"""
        
        # Get tickets assigned to the user
        assigned_tickets = Ticket.objects.filter(
            assigned_to=request.user
        ).exclude(status='closed').count()
        
        # Get resolved/closed tickets
        resolved_tickets = Ticket.objects.filter(
            Q(created_by=request.user) | Q(assigned_to=request.user),
            status='closed'
        ).count()
        
        # Get tasks assigned to the user
        from tenant.models import Task
        assigned_tasks = Task.objects.filter(
            assigned_to=request.user
        ).exclude(task_status='completed').count()
        
        # Get completed tasks
        completed_tasks = Task.objects.filter(
            assigned_to=request.user,
            task_status='completed'
        ).count()
        
        return Response({
            "assigned_tickets": assigned_tickets,
            "resolved_tickets": resolved_tickets,
            "assigned_tasks": assigned_tasks,
            "completed_tasks": completed_tasks,
        }, status=status.HTTP_200_OK)

    def my_customer_dashboard(self, request, *args, **kwargs):
        """List all tickets, optionally without pagination"""
        queryset = Ticket.objects.filter(created_by=request.user, is_merged=False).order_by('-id')

        return Response({
            "all": queryset.count(),
            "pending": queryset.exclude(Q(status="closed") | Q(status="hold")).count(),
            "resolved": queryset.filter(status="closed").count(),
        }, status=status.HTTP_200_OK)

    @transaction.atomic
    def create(self, request, *args, **kwargs):
        try:
            # Extract data from request
            print(request.data)
            
            # Generate ticket ID using business context and config
            ticket_id = Helper().generate_incident_code()

            title = request.data.get('title')
            creator_name = request.data.get('creator_name')
            creator_phone = request.data.get('creator_phone')
            creator_email = request.data.get('creator_email')
            description = request.data.get('description')
            # DB-safe: strip 4-byte characters (e.g. emojis) for MySQL setups that don't support utf8mb4
            description = _sanitize_text_for_mysql(description) if description else description
            category_id = request.data.get('category')
            department_id = request.data.get('department')
            
            priority = request.data.get('priority')
            if not priority:
                priority = 'medium'  # Fallback
                
            customer_tier = request.data.get('customer_tier', 'standard')  # Add customer tier
            source = request.data.get('source', 'web')  # Add source field with default 'web'
            is_public = True if str(request.data.get('is_public')).lower() == 'true' else False
            assignee_id = request.data.get('assignee_id')
            raw_tags = []
            try:
                raw_tags = request.data.getlist('tags[]')
            except Exception:
                pass
            if not raw_tags and 'tags' in request.data:
                maybe_tags = request.data.get('tags')
                if isinstance(maybe_tags, list):
                    raw_tags = maybe_tags
                elif isinstance(maybe_tags, str):
                    raw_tags = [t.strip() for t in maybe_tags.split(',') if t.strip()]
            
            created_by = None

            # Validate required fields
            if not all([title, category_id, department_id, priority]):
                return Response({
                    "message": "Missing required fields: title, category, department, or priority"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validate customer tier
            valid_tiers = ['premium', 'standard', 'basic']
            if customer_tier not in valid_tiers:
                return Response({
                    "message": f"Invalid customer tier. Must be one of: {', '.join(valid_tiers)}"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Get category and department objects
            try:
                category = get_object_or_404(TicketCategories, id=category_id)
            except Http404:
                return Response({
                    "message": f"Category with id {category_id} not found"
                }, status=status.HTTP_404_NOT_FOUND)

            try:
                department = get_object_or_404(Department, id=department_id)
            except Http404:
                return Response({
                    "message": f"Department with id {department_id} not found"
                }, status=status.HTTP_404_NOT_FOUND)

            # Check if user already exists
            if creator_email:
                try:
                    existing_user = Users.objects.filter(
                        email=creator_email
                    ).first()
                    if existing_user:
                        created_by = existing_user
                except Exception as e:
                    # Fallback to create a customer
                    logger.warning(f"Error checking existing user: {str(e)}")

            
            # Optional assignee lookup (must belong to business)
            assigned_agent = None
            if assignee_id not in [None, '', 'null']:
                try:
                    assigned_agent = Users.objects.filter(
                        id=int(assignee_id)
                    ).first()
                    if not assigned_agent:
                        return Response({
                            "message": f"Assignee with id {assignee_id} not found"
                        }, status=status.HTTP_404_NOT_FOUND)
                except ValueError:
                    return Response({
                        "message": "assignee_id must be an integer"
                    }, status=status.HTTP_400_BAD_REQUEST)

            # Normalize tags
            tag_list = []
            for t in raw_tags:
                if isinstance(t, str) and t.strip():
                    tag_list.append(t.strip())
            tag_list = list(dict.fromkeys(tag_list))  # dedupe preserving order

            # Create ticket (without due_date - SLA will calculate it)
            try:
                ticket = Ticket.objects.create(
                    title=title,
                    description=description,
                    category=category,
                    department=department,
                    creator_name=creator_name,
                    creator_email=creator_email,
                    creator_phone=creator_phone,
                    created_by=created_by,
                    ticket_id=ticket_id,
                    priority=priority,
                    customer_tier=customer_tier,  # Add customer tier
                    source=source,  # Add source field
                    is_public=is_public,
                    assigned_to=assigned_agent,
                    tags=",".join(tag_list) if tag_list else ""
                )
                contact = link_or_create_contact(
                    name=creator_name,
                    email=creator_email,
                    phone=creator_phone,
                    owner=request.user,
                )
                if contact:
                    ticket.contact = contact
                    ticket.save(update_fields=["contact"])
                if assigned_agent:
                    ticket.status = 'assigned'
                    ticket.save(update_fields=['status'])


                # Check if SLA is allowed before assigning SLA
                sla_config = SLAConfiguration.objects.filter(pk=1).first()
                sla_enabled = sla_config and sla_config.allow_sla
                
                if sla_enabled:
                    # The SLA object itself will handle the due date calculation
                    applicable_sla = SLA.objects.filter(
                        is_active=True,
                        targets__priority=priority
                    ).first()


                    if applicable_sla:
                        ticket.sla = applicable_sla
                        ticket.save()
                    else:
                        logger.warning(f"No applicable SLA found for ticket {ticket_id} with priority {priority}")

                    # Calculate due dates using the ticket's own methods
                    sla_due_times = ticket.calculate_sla_due_times()
                    due_date = None
                    if sla_due_times and sla_due_times['resolution_due']:
                        due_date = sla_due_times['resolution_due']
                        ticket.due_date = due_date
                        ticket.save()
                        logger.info(f"Ticket {ticket_id} resolution due date set to {due_date} based on SLA.")
                    else:
                        logger.warning(f"Could not calculate SLA due date for ticket {ticket_id}. Falling back to default.")
                        # Fallback to old method if no SLA policy exists or calculation fails
                        try:
                            priority_dict = dict(PRIORITY_DURATION)
                            priority_hours_str = priority_dict.get(priority)
                            if priority_hours_str:
                                priority_hours = int(priority_hours_str)
                                due_date = datetime.now() + timedelta(hours=priority_hours)
                                ticket.due_date = due_date
                                ticket.save()
                            else:
                                due_date = None
                        except Exception as e:
                            logger.error(f"Error with fallback due date calculation: {str(e)}")
                            due_date = None
                else:
                    # SLA is disabled, use fallback method for due date
                    logger.info(f"SLA is disabled, using fallback due date calculation for ticket {ticket_id}")
                    try:
                        priority_dict = dict(PRIORITY_DURATION)
                        priority_hours_str = priority_dict.get(priority)
                        if priority_hours_str:
                            priority_hours = int(priority_hours_str)
                            due_date = datetime.now() + timedelta(hours=priority_hours)
                            ticket.due_date = due_date
                            ticket.save()
                        else:
                            due_date = None
                    except Exception as e:
                        logger.error(f"Error with fallback due date calculation: {str(e)}")
                        due_date = None
                        
            except Exception as e:
                logger.error(f"Error creating ticket or assigning SLA: {str(e)}")
                return Response({
                    "message": "Failed to create ticket or assign SLA",
                    "details": str(e)
                }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            # Optional assignee at creation (internal only) - mirror assign endpoint behavior
            if assignee_id:
                try:
                    agent = get_object_or_404(Users, id=assignee_id)
                except Http404:
                    return Response({
                        "message": f"Assignee with id {assignee_id} not found"
                    }, status=status.HTTP_404_NOT_FOUND)

                old_assignee = ticket.assigned_to
                ticket.assigned_to = agent
                ticket.status = 'assigned'
                ticket.updated_by = request.user
                ticket.updated_at = datetime.now()
                ticket.save(update_fields=['assigned_to', 'status', 'updated_by', 'updated_at'])

                # Activity log
                from tenant.models import TicketActivity
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=request.user,
                    activity_type='assigned',
                    description=f"Ticket Assignment\nTicket was assigned to {agent.get_full_name()}",
                    new_value=agent.get_full_name(),
                    old_value=old_assignee.get_full_name() if old_assignee else ''
                )

                # Notifications
                NotificationSettingsService.create_in_app_notification(
                    user=agent,
                    ticket=ticket,
                    message=f"Ticket #{ticket.ticket_id} has been assigned to you: {ticket.title}",
                    notification_type="ticket_assigned",
                    metadata={
                        "reassigned": bool(old_assignee),
                        "previous_assignee": old_assignee.get_full_name() if old_assignee else None,
                        "priority": ticket.priority,
                        "activity_type": "assigned",
                        "assigned_by": request.user.get_full_name()
                    }
                )
                if old_assignee and old_assignee != agent:
                    NotificationSettingsService.create_in_app_notification(
                        user=old_assignee,
                        ticket=ticket,
                        message=f"Ticket #{ticket.ticket_id} was reassigned from you to {agent.get_full_name()}",
                        notification_type="ticket_assigned",
                        metadata={
                            "reassigned_from": True,
                            "new_assignee": agent.get_full_name(),
                            "priority": ticket.priority,
                            "activity_type": "reassigned"
                        }
                    )

                if not ticket.first_response_at:
                    ticket.mark_first_response()

            # Note: Description is NOT added as a TicketComment to avoid redundancy
            # since it's already displayed in the description section of the ticket

            # Handle file uploads
            if request.FILES:
                logger.info("Files found in request, processing...")
                uploaded_files = request.FILES
                
                for uploaded_file_name, uploaded_file in uploaded_files.items():
                    try:
                        file_extension = os.path.splitext(uploaded_file.name)[1].lower()
                        
                        # Validate file extension (add your allowed extensions)
                        allowed_extensions = ['.jpg', '.jpeg', '.png', '.pdf', '.doc', '.docx', '.txt']
                        if file_extension not in allowed_extensions:
                            return Response({
                                "error": f"File type {file_extension} not allowed"
                            }, status=status.HTTP_400_BAD_REQUEST)
                        
                        # Generate unique filename
                        unique_filename = f"{uuid.uuid4()}{file_extension}"
                        
                        # Create directory (scoped by business)
                        properties_dir, url_base = _build_storage_paths(None, subfolder='files')

                        # Full file path
                        file_path = os.path.join(properties_dir, unique_filename)
                        
                        # Save file
                        with open(file_path, 'wb+') as destination:
                            for chunk in uploaded_file.chunks():
                                destination.write(chunk)
                        
                        # Generate URL
                        file_url = f"{url_base}/{unique_filename}" if url_base else unique_filename
                        
                        # Create attachment
                        TicketAttachment.objects.create(
                            ticket=ticket,
                            file_url=file_url,
                            filename=uploaded_file.name,  # Save original filename
                            description=f"File uploaded for ticket {ticket_id}"
                        )
                        
                        logger.info(f"File saved: {file_path}")
                        
                    except Exception as file_error:
                        logger.error(f"Error saving file {uploaded_file_name}: {str(file_error)}")
                        return Response({
                            "error": f"Failed to save file {uploaded_file_name}",
                            "details": str(file_error)
                        }, status=status.HTTP_400_BAD_REQUEST)

            # Prepare email data
            ticket_email_data = {
                'ticket_id': ticket_id,
                'title': title,
                'due_date': due_date.strftime('%Y-%m-%d %H:%M:%S') if due_date else 'TBD',
                'priority': priority,
                'customer_tier': customer_tier,
                'creator_name': creator_name or 'Anonymous',
                'creator_email': creator_email or 'N/A',
                'creator_phone': creator_phone or 'N/A',
                'category_name': category.name,
                'description': description or 'No description provided',
            }

            # Create in-app notifications for new ticket
            # If ticket has no assignee, notify all department members
            if ticket.assigned_to:
                NotificationSettingsService.create_in_app_notification(
                    user=ticket.assigned_to,
                    ticket=ticket,
                    message=f"New ticket #{ticket.ticket_id} has been assigned to you: {ticket.title}",
                    notification_type="ticket_assigned",
                    metadata={
                        "priority": ticket.priority,
                        "category": category.name,
                        "department": department.name,
                        "activity_type": "assigned"
                    }
                )
            elif department:
                # No assignee - notify all department members
                dept_members = department.get_members()
                NotificationSettingsService.notify_many(
                    users=list(dept_members),
                    ticket=ticket,
                    message=f"New unassigned ticket #{ticket.ticket_id} in {department.name}: {ticket.title}",
                    notification_type="ticket_assigned",
                    metadata={
                        "priority": ticket.priority,
                        "category": category.name,
                        "department": department.name,
                        "unassigned": True,
                        "activity_type": "created"
                    }
                )

            # Prepare response data
            response_data = {
                "message": "Ticket created successfully",
                "ticket_id": ticket_id,
                "priority": priority,
                "customer_tier": customer_tier,
            }
            
            # Add SLA information if available
            if hasattr(ticket, 'sla_tracker'):
                sla_info = ticket.sla_status
                if sla_info:
                    response_data["sla_info"] = {
                        "first_response_due": sla_info['first_response_due'].isoformat(),
                        "resolution_due": sla_info['resolution_due'].isoformat(),
                        "policy_name": ticket.sla_tracker.sla_policy.name
                    }

            # Save Activity
            from tenant.models import TicketActivity
            # Use the actual person who created the ticket (request.user or created_by)
            activity_user = request.user if request.user.is_authenticated else (created_by if created_by else request.user)
            
            # Log ticket creation activity without description (as per user requirement)
            user_name = activity_user.get_full_name() if hasattr(activity_user, 'get_full_name') else str(activity_user)
            TicketActivity.objects.create(
                ticket=ticket,
                user=activity_user,
                activity_type='created',
                description=f"{user_name} created the ticket"
            )

            return Response(response_data, status=status.HTTP_201_CREATED)

        except ValidationError as e:
            logger.error(f"Validation error in ticket creation: {str(e)}")
            return Response({
                "message": "Validation error",
                "details": str(e)
            }, status=status.HTTP_400_BAD_REQUEST)
            
        except Exception as e:
            logger.error(f"Unexpected error in ticket creation: {str(e)}")
            return Response({
                "message": "An unexpected error occurred while creating the ticket",
                "details": str(e) if settings.DEBUG else "Please contact support"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    def add_ticket_watchers(self, request, *args, **kwargs):
        tkId = kwargs.get("id")
        user_ids = request.data.get("watchers", [])  # Expecting a list of user IDs

        if not isinstance(user_ids, list):
            return Response({
                "message": "ids must be a list of user IDs."
            }, status=status.HTTP_400_BAD_REQUEST)

        ticket = get_object_or_404(Ticket, id=tkId)
        if ticket.status == 'closed':
            return Response({"message": "Cannot add watchers to a closed ticket."}, status=status.HTTP_400_BAD_REQUEST)

        added = []
        skipped = []
        errors = []

        for user_id in user_ids:
            try:
                user = Users.objects.filter(id=user_id).first()

                # Check if user is already a watcher
                if TicketWatchers.objects.filter(ticket=ticket, watcher=user).exists():
                    skipped.append(user_id)
                    continue

                # Create the watcher entry
                TicketWatchers.objects.create(ticket=ticket, watcher=user)
                added.append(user_id)
                
                # Log activity for watcher addition
                from tenant.models import TicketActivity
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=request.user,
                    activity_type='watcher_added',
                    description=f"Watcher Added\n{request.user.get_full_name()} added {user.get_full_name()} as a watcher",
                    old_value='',
                    new_value=user.get_full_name()
                )
                
                # Notification is now handled by signals

            except Users.DoesNotExist:
                errors.append({"id": user_id, "error": "User not found"})
            except IntegrityError as e:
                errors.append({"id": user_id, "error": str(e)})
            except Exception as e:
                errors.append({"id": user_id, "error": str(e)})

        return Response({
            "message": "Watcher update completed.",
            "added": added,
            "skipped": skipped,
            "errors": errors
        }, status=status.HTTP_200_OK)

    def get_watchers(request, *args, **kwargs):
        tkId = kwargs.get("id")
        ticket = get_object_or_404(Ticket, id=tkId)
        watchers = TicketWatchers.objects.filter(ticket=ticket).select_related("watcher")
        watcher_list = [
            {
                "id": w.watcher.id,
                "name": w.watcher.full_name(),
                "email": w.watcher.email,
                "avatar_url": w.watcher.avatar_url,
            }
            for w in watchers
        ]
        return Response(watcher_list, status=status.HTTP_200_OK)

    @transaction.atomic
    def add_ticket_tags(self, request, *args, **kwargs):
        tkId = kwargs.get("id")
        tags = request.data.get("tags", [])

        if not isinstance(tags, list):
            return Response({
                "message": "Tags must be a list of strings."
            }, status=status.HTTP_400_BAD_REQUEST)

        ticket = get_object_or_404(Ticket, id=tkId)
        if ticket.status == 'closed':
            return Response({"message": "Cannot add tags to a closed ticket."}, status=status.HTTP_400_BAD_REQUEST)

        # Get existing tags from DB and normalize
        existing_tags = [t.strip() for t in ticket.tags.split(",") if t.strip()]
        new_tags = [t.strip() for t in tags if t.strip()]

        new_set = set(new_tags)


        # Final set of tags
        final_tags = list(new_set)  # overwrite with new ones only

        # Save back as comma-separated string
        ticket.tags = ",".join(final_tags)
        ticket.save()
        
        # Log activity for each new tag added
        from tenant.models import TicketActivity
        newly_added = [tag for tag in new_tags if tag not in existing_tags]
        for tag in newly_added:
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='tag_added',
                description=f"Tag Added\n{request.user.get_full_name()} added tag '{tag}'",
                old_value='',
                new_value=tag
            )

        return Response({
            "message": "Tags update completed.",
        }, status=status.HTTP_200_OK)

    def get_tags(self, request, *args, **kwargs):
        tkId = kwargs.get("id")
        ticket = get_object_or_404(Ticket, id=tkId)

        # Split comma-separated tags into list
        tags = [t.strip() for t in ticket.tags.split(",") if t.strip()]

        return Response({
            "tags": tags,
        }, status=status.HTTP_200_OK)

    @transaction.atomic
    def assign(self, request, *args, **kwargs):
        try:
            ticket_id = request.data.get('ticket_id')
            agent_id = request.data.get('agent_id')

            if not ticket_id:
                return Response({
                    "message": "Ticket ID is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            ticket = get_object_or_404(Ticket, id=ticket_id)
            if ticket.status == 'closed':
                return Response({"message": "Cannot assign a closed ticket."}, status=status.HTTP_400_BAD_REQUEST)
            
            old_assignee = ticket.assigned_to
            
            # Handle unassignment when agent_id is None or null
            if agent_id is None:
                ticket.assigned_to = None
                ticket.updated_at = datetime.now()
                ticket.status = 'created'
                ticket.updated_by = request.user
                ticket.save()

                # Create TicketActivity for unassignment
                from tenant.models import TicketActivity
                
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=request.user,
                    activity_type='unassigned',
                    description=f"Ticket Unassignment\nTicket was unassigned from {old_assignee.get_full_name() if old_assignee else 'agent'}",
                    old_value=old_assignee.get_full_name() if old_assignee else '',
                    new_value=''  # Empty string instead of None to avoid NULL constraint error
                )

                # Notify the old assignee
                if old_assignee:
                    NotificationSettingsService.create_in_app_notification(
                        user=old_assignee,
                        ticket=ticket,
                        message=f"Ticket #{ticket.ticket_id} was unassigned from you",
                        notification_type="ticket_assigned",
                        metadata={
                            "unassigned": True,
                            "priority": ticket.priority,
                            "activity_type": "unassigned"
                        }
                    )
                
                # Notify department members when ticket is unassigned
                if ticket.department:
                    dept_members = ticket.department.get_members()
                    if old_assignee:
                        dept_members = dept_members.exclude(id=old_assignee.id)
                    
                    NotificationSettingsService.notify_many(
                        users=list(dept_members),
                        ticket=ticket,
                        message=f"Ticket #{ticket.ticket_id} is now unassigned and needs attention",
                        notification_type="ticket_assigned",
                        metadata={
                            "unassigned": True,
                            "department": ticket.department.name,
                            "priority": ticket.priority,
                            "activity_type": "unassigned"
                        }
                    )

                return Response({
                    "message": "Ticket unassigned successfully"
                }, status=status.HTTP_200_OK)
            
            # Handle assignment to a specific agent
            if not agent_id:
                return Response({
                    "message": "Agent ID is required for assignment"
                }, status=status.HTTP_400_BAD_REQUEST)
                
            agent = get_object_or_404(Users, id=agent_id)
            ticket.assigned_to = agent
            ticket.updated_at = datetime.now()
            ticket.status = 'assigned'
            ticket.updated_by = request.user
            ticket.save()

            # Create TicketActivity for assignment
            from tenant.models import TicketActivity
            
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='assigned',
                description=f"Ticket Assignment\nTicket was assigned to {agent.get_full_name()}",
                new_value=agent.get_full_name(),
                old_value=old_assignee.get_full_name() if old_assignee else ''  # Empty string instead of None
            )

            # Create in-app notifications for assignment
            # Notify new assignee
            if agent:
                NotificationSettingsService.create_in_app_notification(
                    user=agent,
                    ticket=ticket,
                    message=f"Ticket #{ticket.ticket_id} has been assigned to you: {ticket.title}",
                    notification_type="ticket_assigned",
                    metadata={
                        "reassigned": bool(old_assignee),
                        "previous_assignee": old_assignee.get_full_name() if old_assignee else None,
                        "priority": ticket.priority,
                        "activity_type": "assigned",
                        "assigned_by": request.user.get_full_name()
                    }
                )
            
            # Notify previous assignee if they exist
            if old_assignee and old_assignee != agent:
                new_assignee_name = agent.get_full_name() if agent else "Unassigned"
                NotificationSettingsService.create_in_app_notification(
                    user=old_assignee,
                    ticket=ticket,
                    message=f"Ticket #{ticket.ticket_id} was reassigned from you to {new_assignee_name}",
                    notification_type="ticket_assigned",
                    metadata={
                        "reassigned_from": True,
                        "new_assignee": new_assignee_name,
                        "priority": ticket.priority,
                        "activity_type": "reassigned"
                    }
                )

            # Mark first response if not already marked
            if not ticket.first_response_at:
                ticket.mark_first_response()

            return Response({
                "message": "Ticket assigned successfully"
            }, status=status.HTTP_200_OK)

        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

        except Users.DoesNotExist:
            logger.warning(f"Agent with ID {agent_id} not found")
            return Response({"message": "Agent not found"}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            logger.exception("An error occurred while assigning the ticket.", exc_info=e)
            return Response({
                "message": "An unexpected error occurred. Please try again later."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    def assign_to_me(self, request, *args, **kwargs):
        try:
            ticket_id = kwargs.get('id')

            if not ticket_id:
                return Response({
                    "message": "Ticket ID are required"
                }, status=status.HTTP_400_BAD_REQUEST)

            ticket = get_object_or_404(Ticket, id=ticket_id)

            ticket.assigned_to = request.user
            ticket.updated_at = datetime.now()
            ticket.status = 'assigned'
            ticket.updated_by = request.user
            ticket.save()

            # Create TicketActivity for claim
            from tenant.models import TicketActivity
            
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='assigned',
                description=f"Ticket Assignment\nTicket was claimed by {request.user.get_full_name()}",
                new_value=request.user.get_full_name()
            )

            # The claim task is specific and not easily handled by a generic signal
            transaction.on_commit(lambda: ticket_claim_task.delay(ticket.id))

            # Send notification to the agent
            # send_ticket_assignment_email.delay(ticket.id, agent.email)
            # notification_message = f"You have been assigned to a new ticket #{ticket.ticket_id}: {ticket.title}"
            # create_notification_task.delay(
            #     user_id=ticket.assigned_to.id,
            #     ticket_id=ticket.id,
            #     message=notification_message,
            #     notification_type="ticket_assigned",
            # )
            # if old_assignee and old_assignee != agent:
            #     reassign_message = f"Ticket #{ticket.ticket_id} was reassigned from you to {agent.get_full_name()}"
            #     create_notification_task.delay(
            #         user_id=old_assignee.id,
            #         ticket_id=ticket.id,
            #         message=reassign_message,
            #         notification_type="ticket_assigned",
            #     )

            # Mark fisrt response if not already marked
            if not ticket.first_response_at:
                ticket.mark_first_response()

            return Response({
                "message": "Ticket claimed successfully"
            }, status=status.HTTP_200_OK)

        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

        except Exception as e:
            logger.exception("An error occurred while assigning the ticket.", exc_info=e)
            return Response({
                "message": "An unexpected error occurred. Please try again later."
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)



    @transaction.atomic
    def update_status(self, request, *args, **kwargs):
        ticket_id = kwargs.get('id')

        if not ticket_id:
            return Response({
                "message": "Ticket ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        

        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            if ticket.status == 'closed':
                return Response({"message": "Cannot update status of a closed ticket. Please reopen it first."}, status=status.HTTP_400_BAD_REQUEST)
            new_status = request.data.get("status")
            old_status = ticket.status
            notes = request.data.get('notes')

            if not new_status:
                return Response({
                    "message": "Status is required"
                }, status=status.HTTP_400_BAD_REQUEST)
            
            # Prevent closing ticket if there are incomplete tasks
            if new_status in ['closed', 'resolved']:
                incomplete_tasks = Task.objects.filter(
                    linked_ticket=ticket,
                    is_archived=False,
                    is_deleted=False
                ).exclude(task_status__in=['completed', 'cancelled'])
                
                if incomplete_tasks.exists():
                    incomplete_count = incomplete_tasks.count()
                    return Response({
                        "message": f"Cannot close ticket. There {'is' if incomplete_count == 1 else 'are'} {incomplete_count} incomplete task{'s' if incomplete_count > 1 else ''} linked to this ticket. Please complete or cancel all tasks before closing the ticket."
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Allow customers/users to close unassigned tickets
            # Only agents need assignment restrictions for other status changes
            if not ticket.assigned_to and new_status not in ['closed']:
                return Response({
                    "message": "Please assign this ticket before changing its status. Use 'Assign to me' to take ownership."
                }, status=status.HTTP_403_FORBIDDEN)

            if new_status == 'on_hold':
                ticket.pause_sla(reason=notes)
            elif old_status == 'on_hold' and new_status != 'on_hold':
                ticket.resume_sla()

            # Note: 'resolved' is now a valid status, don't auto-close
            # Only set resolved_at when resolving
            if new_status == 'resolved' and old_status != 'resolved':
                ticket.resolved_at = timezone.now()

            # Update the ticket status
            ticket.status = new_status
            ticket.notes = notes
            ticket.updated_at = datetime.now()
            ticket.updated_by = request.user
            ticket.save()

            # Create a TicketComment for notes if provided
            if notes:
                from tenant.models import TicketComment
                TicketComment.objects.create(
                    ticket=ticket,
                    author=request.user,
                    content=f"**Status Update Note**\n\n{notes}",
                    is_internal=False
                )

            # Create a TicketActivity for the status update
            from tenant.models import TicketActivity
            
            # Get display name for the status
            status_display = dict(Ticket.STATUS_CHOICES).get(new_status, new_status.replace('_', ' ').title())
            user_name = request.user.get_full_name() or request.user.email
            
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='status_changed',
                description=f"{user_name} changed the ticket status to {status_display}",
                old_value=old_status,
                new_value=new_status
            )

            # Create in-app notification for status change
            if ticket.assigned_to:
                status_message = f"Ticket #{ticket.ticket_id} status changed from {old_status.title()} to {new_status.title()}"
                
                # Special handling for reopened tickets
                if old_status in ["resolved", "closed"] and new_status in ["open", "in_progress"]:
                    notification_type = "ticket_reopened"
                    status_message = f"Ticket #{ticket.ticket_id} has been reopened"
                else:
                    notification_type = "ticket_status_changed"
                
                NotificationSettingsService.create_in_app_notification(
                    user=ticket.assigned_to,
                    ticket=ticket,
                    message=status_message,
                    notification_type=notification_type,
                    metadata={
                        "old_status": old_status,
                        "new_status": new_status,
                        "priority": ticket.priority,
                        "activity_type": "status_changed",
                        "changed_by": request.user.get_full_name()
                    }
                )
            
            return Response({
                "message": "Ticket status updated successfully"
            }, status=status.HTTP_200_OK)
        
        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    def update_department(self, request, *args, **kwargs):
        ticket_id = kwargs.get('id')

        if not ticket_id:
            return Response({
                "message": "Ticket ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            
            # Prevent modification of closed tickets
            if ticket.status == 'closed':
                return Response(
                    {"message": "Cannot modify a closed ticket. Please reopen it first."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            department_id = request.data.get("department_id")
            old_department = ticket.department

            if not department_id:
                return Response({
                    "message": "Department ID is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Verify the department exists and belongs to the business
            from tenant.models import Department
            try:
                department = Department.objects.get(id=department_id)
            except Department.DoesNotExist:
                return Response({
                    "message": "Department not found"
                }, status=status.HTTP_404_NOT_FOUND)

            # Update the ticket department
            ticket.department = department
            ticket.updated_at = datetime.now()
            ticket.updated_by = request.user
            ticket.save()

            # Create a TicketActivity for the department change
            from tenant.models import TicketActivity
            
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='department_changed',
                description=f"Ticket department changed from {old_department.name if old_department else 'None'} to {department.name}",
                old_value=old_department.name if old_department else 'None',
                new_value=department.name
            )

            return Response({
                "message": "Ticket department updated successfully"
            }, status=status.HTTP_200_OK)

        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    def update_category(self, request, *args, **kwargs):
        ticket_id = kwargs.get('id')

        if not ticket_id:
            return Response({
                "message": "Ticket ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            
            # Prevent modification of closed tickets
            if ticket.status == 'closed':
                return Response(
                    {"message": "Cannot modify a closed ticket. Please reopen it first."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            category_id = request.data.get("category_id")
            old_category = ticket.category

            if not category_id:
                return Response({
                    "message": "Category ID is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Verify the category exists and belongs to the business context
            from tenant.models import TicketCategories
            try:
                category = TicketCategories.objects.get(id=category_id)
            except TicketCategories.DoesNotExist:
                return Response({
                    "message": "Category not found"
                }, status=status.HTTP_404_NOT_FOUND)

            # Update the ticket category
            ticket.category = category
            ticket.updated_at = datetime.now()
            ticket.updated_by = request.user
            ticket.save()

            # Create a TicketActivity for the category change
            from tenant.models import TicketActivity
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='category_changed',
                description=f"Ticket category changed from {old_category.name if old_category else 'None'} to {category.name}",
                old_value=old_category.name if old_category else 'None',
                new_value=category.name
            )

            return Response({
                "message": "Ticket category updated successfully"
            }, status=status.HTTP_200_OK)

        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    def update_priority(self, request, *args, **kwargs):
        ticket_id = kwargs.get('id')

        if not ticket_id:
            return Response({
                "message": "Ticket ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            
            # Prevent modification of closed tickets
            if ticket.status == 'closed':
                return Response(
                    {"message": "Cannot modify a closed ticket. Please reopen it first."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            new_priority = request.data.get('priority')
            old_priority = ticket.priority

            if not new_priority:
                return Response({"message": "Priority is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Validate priority value (optional: ensure it's one of allowed choices)
            ticket.priority = new_priority
            ticket.updated_at = datetime.now()
            ticket.updated_by = request.user
            ticket.save()

            # Create TicketActivity for priority change
            from tenant.models import TicketActivity
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='priority_changed',
                description=f"Ticket priority changed from {old_priority} to {new_priority}",
                old_value=old_priority,
                new_value=new_priority
            )

            return Response({"message": "Ticket priority updated successfully"}, status=status.HTTP_200_OK)

        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    def update_source(self, request, *args, **kwargs):
        """Update the source of a ticket (e.g., email, phone, web, chat)."""
        ticket_id = kwargs.get('id')

        if not ticket_id:
            return Response({
                "message": "Ticket ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            new_source = request.data.get('source')
            old_source = ticket.source

            if not new_source:
                return Response({"message": "Source is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Valid source options - matches frontend options
            valid_sources = ['email', 'phone', 'web', 'chat', 'chatbot', 'api', 'internal', 'customer_portal', 'portal', 'other']
            if new_source.lower() not in valid_sources:
                return Response({
                    "message": f"Invalid source. Valid options are: {', '.join(valid_sources)}"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Update the ticket source
            ticket.source = new_source.lower()
            ticket.updated_at = datetime.now()
            ticket.updated_by = request.user
            ticket.save()

            # Create TicketActivity for source change
            from tenant.models import TicketActivity
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='source_changed',
                description=f"Ticket source changed from {old_source or 'None'} to {new_source}",
                old_value=old_source or 'None',
                new_value=new_source
            )

            return Response({"message": "Ticket source updated successfully"}, status=status.HTTP_200_OK)

        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    def update_due_date(self, request, *args, **kwargs):
        """Update the due date of a ticket."""
        ticket_id = kwargs.get('id')

        if not ticket_id:
            return Response({
                "message": "Ticket ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)

        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            
            # Prevent modification of closed tickets
            if ticket.status == 'closed':
                return Response(
                    {"message": "Cannot modify a closed ticket. Please reopen it first."},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            new_due_date = request.data.get('due_date')
            old_due_date = ticket.due_date

            if not new_due_date:
                return Response({"message": "Due date is required"}, status=status.HTTP_400_BAD_REQUEST)

            try:
                # Parse the due date string to datetime
                # Expected format: ISO 8601 (e.g., "2024-12-31T23:59:59Z" or "2024-12-31")
                if isinstance(new_due_date, str):
                    # Handle both date and datetime strings
                    if 'T' in new_due_date:
                        parsed_date = datetime.fromisoformat(new_due_date.replace('Z', '+00:00'))
                    else:
                        parsed_date = datetime.strptime(new_due_date, '%Y-%m-%d')
                        # Set to end of day if only date provided
                        parsed_date = parsed_date.replace(hour=23, minute=59, second=59)
                    
                    # Make timezone aware if it's not already
                    if timezone.is_naive(parsed_date):
                        parsed_date = timezone.make_aware(parsed_date)
                    
                    new_due_date = parsed_date
                else:
                    new_due_date = timezone.make_aware(new_due_date) if timezone.is_naive(new_due_date) else new_due_date

            except (ValueError, TypeError) as e:
                return Response({
                    "message": f"Invalid due date format. Expected ISO 8601 format (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS). Error: {str(e)}"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Validate that due date is in the future
            if new_due_date <= timezone.now():
                return Response({
                    "message": "Due date must be in the future"
                }, status=status.HTTP_400_BAD_REQUEST)

            # Update the ticket due date
            ticket.due_date = new_due_date
            ticket.updated_at = datetime.now()
            ticket.updated_by = request.user
            ticket.save()

            # Create TicketActivity for due date change
            from tenant.models import TicketActivity
            old_due_date_str = old_due_date.strftime('%Y-%m-%d %H:%M:%S') if old_due_date else 'None'
            new_due_date_str = new_due_date.strftime('%Y-%m-%d %H:%M:%S')
            
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='due_date_changed',
                description=f"Ticket due date changed from {old_due_date_str} to {new_due_date_str}",
                old_value=old_due_date_str,
                new_value=new_due_date_str
            )

            return Response({
                "message": "Ticket due date updated successfully",
                "due_date": new_due_date.isoformat()
            }, status=status.HTTP_200_OK)

        except Ticket.DoesNotExist:
            logger.warning(f"Ticket with ID {ticket_id} not found")
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    @transaction.atomic
    @action(detail=True, methods=['post'], url_path='add-note')
    def add_note(self, request, *args, **kwargs):
        ticket_id = kwargs.get('pk')
        if not ticket_id:
            return Response({"message": "Ticket ID is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            note = request.data.get('note')

            if not note:
                return Response({"message": "Note content is required"}, status=status.HTTP_400_BAD_REQUEST)

            # Notes are internal comments
            ticket.comments.create(
                ticket=ticket,
                author=request.user,
                content=note,
                updated_by=request.user,
                is_internal=True
            )

            return Response({"message": "Note added successfully"}, status=status.HTTP_201_CREATED)
        
        except Ticket.DoesNotExist:
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    @transaction.atomic    
    def add_comment(self, request, *args, **kwargs):
        ticket_id = kwargs.get('id')
        if not ticket_id:
            return Response({
                "message": "Ticket ID is required"
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)
            comment = request.data.get('comment')
            # Sanitize comment for MySQL if necessary
            comment = _sanitize_text_for_mysql(comment) if comment else comment
            is_internal = str(request.data.get("is_internal", "false")).lower() == "true"
            # logger.info(f"Adding comment by user: {request.user} (ID: {request.user.id}, Email: {request.user.email})")

            # Set author if user is authenticated, otherwise None
            author = request.user if request.user and request.user.is_authenticated else None
            print(f"Author set to: {author}")
            # Set updated_by: author if present, otherwise system user
            if author:
                updated_by = author
            else:
                updated_by = Users.objects.filter(first_name="System").first()

            com = ticket.comments.create(
                ticket=ticket,
                author=author,
                content=comment,
                updated_by=updated_by,
                is_internal=is_internal
            )

            # ========== EXTRACT AND SAVE BASE64 IMAGES ==========
            # Extract embedded base64 images from HTML content and save as attachments
            try:
                if comment and '<img' in comment and 'base64' in comment:
                    updated_content = _extract_and_save_base64_images(
                        html_content=comment,
                        comment=com,
                        business=ticket.business
                    )
                    # Update comment content with file URLs instead of base64
                    if updated_content != comment:
                        com.content = updated_content
                        com.save(update_fields=['content'])
            except Exception as img_error:
                logger.error(f"Failed to extract base64 images from comment: {img_error}")
            # =====================================================

            # ========== SET HAS_NEW_REPLY BADGE ==========
            # If the comment is from someone other than the assigned agent, set has_new_reply
            try:
                if not is_internal:  # Only for non-internal (public) comments
                    # If author is None (anonymous/customer) or author is not the assigned agent
                    if author is None or (ticket.assigned_to_id and author.id != ticket.assigned_to_id):
                        ticket.has_new_reply = True
                        ticket.save(update_fields=['has_new_reply'])
            except Exception as badge_error:
                logger.error(f"Failed to set has_new_reply badge: {badge_error}")
            # =====================================================

            # ========== AUTO STATUS TRANSITION ==========
            # When agent replies to ticket in 'open' status, auto-change to 'in_progress'
            try:
                if author and ticket.status == 'open':
                    old_status = ticket.status
                    ticket.status = 'in_progress'
                    ticket.updated_at = datetime.now()
                    ticket.updated_by = author
                    ticket.save(update_fields=['status', 'updated_at', 'updated_by'])
                    
                    # Log the auto-transition in activity stream
                    from tenant.models import TicketActivity
                    TicketActivity.objects.create(
                        ticket=ticket,
                        user=author,
                        activity_type='status_changed',
                        description="System changed the ticket status to In Progress",
                        old_value=old_status,
                        new_value='in_progress'
                    )
                    logger.info(f"Auto-transitioned ticket {ticket.ticket_id} from 'open' to 'in_progress' on first agent reply")
            except Exception as auto_status_error:
                logger.error(f"Failed to auto-transition ticket status: {auto_status_error}")
            # =====================================================

            # ========== DETECT @ MENTIONS AND CREATE NOTIFICATIONS ==========
            # Wrapped in try-except to prevent comment creation from failing if notification fails
            try:
                # Detect @ mentions and create notifications
                if comment and author:
                    # Extract all @ mentions from the comment
                    mentions = re.findall(r'@([A-Za-z]+(?:\s+[A-Za-z]+)*)', comment)
                    
                    for mention_name in mentions:
                        mention_name = mention_name.strip()
                        if not mention_name:
                            continue
                        
                        # Try to find user by full name, first name, or last name
                        mentioned_users = Users.objects.filter(
                            Q(full_name__icontains=mention_name) |
                            Q(first_name__icontains=mention_name) |
                            Q(last_name__icontains=mention_name)
                        )
                        
                        # Create notification for each matched user
                        for mentioned_user in mentioned_users:
                            # Don't notify the author about their own mention
                            if mentioned_user.id != author.id:
                                try:
                                    # Resolve a human-friendly display name for the author
                                    if hasattr(author, "full_name") and callable(author.full_name):
                                        author_name = author.full_name()
                                    else:
                                        author_name = getattr(author, "email", str(author))

                                    NotificationSettingsService.create_in_app_notification(
                                        user=mentioned_user,
                                        ticket=ticket,
                                        message=f"{author_name} mentioned you in ticket #{ticket.ticket_id}",
                                        notification_type='ticket_mention',
                                        metadata={
                                            # Backwards-compatible keys
                                            'comment_id': com.id,
                                            'mentioned_by': author.id,
                                            'mentioned_by_name': author_name,
                                            # Normalised keys for new consumers
                                            'mentioned_by_id': author.id,
                                            'mentioned_by_display': author_name,
                                        }
                                    )
                                    
                                    # ========== Send Email Notification (async) ==========
                                    try:
                                        send_mention_email_notification.delay(
                                            mentioned_user_id=mentioned_user.id,
                                            ticket_id=ticket.id,
                                            mentioned_by_id=author.id,
                                            comment_text=comment
                                        )
                                    except Exception as email_error:
                                        logger.error(f"Failed to queue mention email: {email_error}")
                                    # ====================================================
                                    
                                except Exception as notif_error:
                                    logger.error(f"Failed to create mention notification: {notif_error}")
            
            except Exception as mention_error:
                # Log the error but don't fail the comment creation
                logger.error(f"Failed to process @ mentions, but comment was saved: {mention_error}")
                import traceback
                logger.error(traceback.format_exc())
            # ===================================================================

            # ========== CREATE REGULAR COMMENT NOTIFICATIONS ==========
            # Notify assigned agent, ticket creator, watchers, and department members
            # Using NotificationSettingsService for proper in-app notifications
            try:
                if author:
                    recipients = []
                    
                    # Notify assigned agent (if exists and not author)
                    if ticket.assigned_to and ticket.assigned_to != author:
                        recipients.append(ticket.assigned_to)
                    
                    # If no assignee and ticket has department, notify all department members (admins)
                    if not ticket.assigned_to and ticket.department:
                        dept_members = ticket.department.get_members().exclude(id=author.id)
                        recipients.extend(list(dept_members))
                    
                    # Notify ticket creator (if not author/assignee and not internal note)
                    if (not is_internal and ticket.created_by and 
                        ticket.created_by != author and 
                        ticket.created_by != ticket.assigned_to):
                        recipients.append(ticket.created_by)
                    
                    # Notify watchers (if not author)
                    for watcher in ticket.watchers.all():
                        if watcher.watcher and watcher.watcher != author:
                            recipients.append(watcher.watcher)
                    
                    # Remove duplicates while preserving order
                    seen = set()
                    unique_recipients = []
                    for recipient in recipients:
                        if recipient.id not in seen:
                            seen.add(recipient.id)
                            unique_recipients.append(recipient)
                    
                    # Create notifications for all recipients
                    comment_preview = comment[:100] + "..." if len(comment) > 100 else comment
                    
                    message = (
                        f"New internal note added to ticket #{ticket.ticket_id}"
                        if is_internal else
                        f"New comment on ticket #{ticket.ticket_id}: {comment_preview}"
                    )
                    
                    # Resolve a consistent author name/id for metadata
                    if author:
                        if hasattr(author, "full_name") and callable(author.full_name):
                            author_name = author.full_name()
                        else:
                            author_name = getattr(author, "email", str(author))
                        author_id = author.id
                    else:
                        author_name = "Anonymous"
                        author_id = None

                    NotificationSettingsService.notify_many(
                        users=unique_recipients,
                        ticket=ticket,
                        message=message,
                        notification_type="ticket_comment",
                        metadata={
                            "comment_id": com.id,
                            # Backwards-compatible key
                            "author": author_name,
                            # Normalised keys for new consumers
                            "author_id": author_id,
                            "author_name": author_name,
                            "is_internal": is_internal,
                            "activity_type": "comment",
                        }
                    )
                            
            except Exception as comment_notif_error:
                # Log the error but don't fail the comment creation
                logger.error(f"Failed to process regular comment notifications: {comment_notif_error}")
                import traceback
                logger.error(traceback.format_exc())
            # ==========================================================

            if request.FILES:
                uploaded_files = request.FILES

                for uploaded_file_name, uploaded_file in uploaded_files.items():
                    try:
                        file_extension = os.path.splitext(uploaded_file.name)[1].lower()

                        # Generate unique filename
                        unique_filename = f"{uuid.uuid4()}{file_extension}"

                        # Create directory (scoped by business)
                        properties_dir, url_base = _build_storage_paths(None, subfolder='files')

                        # Full file path
                        file_path = os.path.join(properties_dir, unique_filename)

                        # Save file
                        with open(file_path, 'wb+') as destination:
                            for chunk in uploaded_file.chunks():
                                destination.write(chunk)

                        # Generate URL (be defensive: FILE_URL may be empty in some envs)
                        file_url_base = url_base or FILE_URL or getattr(settings, 'FILE_BASE_URL', '')
                        if file_url_base:
                            file_url = f"{file_url_base.rstrip('/')}/{unique_filename}"
                        else:
                            # fallback to relative path if no base URL configured
                            file_url = unique_filename

                        # Create attachment
                        TicketReplayAttachment.objects.create(
                            comment=com,
                            file_url=file_url,
                            filename=uploaded_file.name  # Save original filename
                        )

                    except Exception as file_error:
                        logger.exception("Error saving uploaded file %s", uploaded_file_name)
                        return Response({
                            "message": f"Failed to save file {uploaded_file_name}",
                            "details": str(file_error)
                        }, status=status.HTTP_400_BAD_REQUEST)

            return Response({
                "message": "Comment added successfully"
            }, status=status.HTTP_201_CREATED)
        
        except Ticket.DoesNotExist:
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception("Unhandled error in add_comment: %s", e)
            return Response({
                "message": "An unexpected error occurred while adding the comment",
                "details": str(e) if settings.DEBUG else "Please contact support"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    @action(detail=True, methods=['post'], url_path='email-reply')
    def send_email_reply(self, request, pk=None, *args, **kwargs):
        """
        Send an email reply for a ticket from the activity stream.
        
        Request body:
        {
            "to": ["recipient@example.com"],
            "cc": ["cc@example.com"],  # optional
            "bcc": ["bcc@example.com"],  # optional
            "content": "Hi John,\n\nThank you for reaching out...",
            "mailbox_id": 5,  # optional
            "close_ticket": false  # optional
        }
        """
        if not pk:
            return Response({"message": "Ticket ID is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            ticket = get_object_or_404(Ticket, id=pk)
            
            # Validate recipients
            to_emails = request.data.get('to', [])
            if not to_emails or not isinstance(to_emails, list):
                return Response({"message": "'to' must be a non-empty list of emails"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Basic email validation
            import re
            email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
            for email in to_emails:
                if not re.match(email_pattern, email):
                    return Response({"message": f"Invalid email address: {email}"}, status=status.HTTP_400_BAD_REQUEST)
            
            cc_emails = request.data.get('cc', []) or []
            bcc_emails = request.data.get('bcc', []) or []
            content = request.data.get('content', '')
            mailbox_id = request.data.get('mailbox_id')
            close_ticket = request.data.get('close_ticket', False)
            
            if not content.strip():
                return Response({"message": "Email content is required"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Sanitize content
            content = _sanitize_text_for_mysql(content) if content else content
            
            # Create comment record (non-internal since it's going to customer)
            author = request.user
            com = ticket.comments.create(
                ticket=ticket,
                author=author,
                content=content,
                updated_by=author,
                is_internal=False,
                email_to=to_emails,
                email_cc=cc_emails if cc_emails else None,
                email_bcc=bcc_emails if bcc_emails else None
            )
            
            # ========== EXTRACT AND SAVE BASE64 IMAGES ==========
            # Extract embedded base64 images from HTML content and save as attachments
            try:
                if content and '<img' in content and 'base64' in content:
                    updated_content = _extract_and_save_base64_images(
                        html_content=content,
                        comment=com,
                        business=ticket.business
                    )
                    # Update comment content with file URLs instead of base64
                    if updated_content != content:
                        com.content = updated_content
                        com.save(update_fields=['content'])
                        content = updated_content  # Use updated content for email
            except Exception as img_error:
                logger.error(f"Failed to extract base64 images from email reply: {img_error}")
            # =====================================================
            
            # Queue email sending via Celery
            from shared.workers.email_reply import send_email_reply_task
            send_email_reply_task.delay(
                ticket_id=ticket.id,
                comment_id=com.id,
                to=to_emails,
                cc=cc_emails if cc_emails else None,
                bcc=bcc_emails if bcc_emails else None,
                mailbox_id=mailbox_id,
            )
            
            # Optionally close ticket after sending
            if close_ticket:
                ticket.status = 'closed'
                ticket.resolved_at = timezone.now()
                ticket.updated_by = author
                ticket.save()
                
                # Create activity for close
                from tenant.models import TicketActivity
                TicketActivity.objects.create(
                    ticket=ticket,
                    user=author,
                    activity_type='status_changed',
                    description="Ticket closed after sending email reply",
                    old_value=ticket.status,
                    new_value='closed'
                )
            
            return Response({
                "message": "Email reply queued successfully",
                "comment_id": com.id,
            }, status=status.HTTP_201_CREATED)
        
        except Ticket.DoesNotExist:
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception("Error in send_email_reply: %s", e)
            return Response({
                "message": "Failed to send email reply",
                "details": str(e) if settings.DEBUG else "Please try again"
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    @action(detail=True, methods=['post'])
    def reopen(self, request, *args, **kwargs):
        ticket_id = kwargs.get('pk')
        reason = request.data.get('reason')

        if not ticket_id:
            return Response({"message": "Ticket ID is required"}, status=status.HTTP_400_BAD_REQUEST)
        
        if not reason:
            return Response({"message": "Reason is required"}, status=status.HTTP_400_BAD_REQUEST)

        try:
            ticket = get_object_or_404(Ticket, id=ticket_id)

            if ticket.status != 'closed':
                return Response({"message": "Only closed tickets can be reopened."}, status=status.HTTP_400_BAD_REQUEST)

            # Reopen ticket
            old_status = ticket.status
            ticket.status = 'open'
            ticket.resolved_at = None
            ticket.updated_at = timezone.now()
            ticket.updated_by = request.user
            ticket.save()

            # Create a record of the reopen event
            TicketReopen.objects.create(
                ticket=ticket,
                reopened_by=request.user,
                reason=reason
            )

            # Create a TicketComment for reopen action (so it appears in activity stream)
            from tenant.models import TicketComment
            TicketComment.objects.create(
                ticket=ticket,
                author=request.user,
                content=f"**Ticket Reopened**\n\nReason: {reason}",
                is_internal=False
            )

            # Create TicketActivity for reopen action
            from tenant.models import TicketActivity
            
            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='status_changed',
                description=f"Ticket has been reopened by {request.user.get_full_name()}\nReason: {reason}",
                old_value=old_status,
                new_value='open'
            )

            # NOTIFICATIONS ARE NOW HANDLED BY SIGNAL HANDLER in shared/signals/notifications.py
            # The signal automatically creates notification for ticket reopened
            # DO NOT create notifications here - it will create duplicates!

            return Response({"message": "Ticket reopened successfully"}, status=status.HTTP_200_OK)
        
        except Ticket.DoesNotExist:
            return Response({"message": "Ticket not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='comments/(?P<comment_id>\d+)/flag')
    def flag_comment(self, request, pk=None, comment_id=None):
        """Flag or unflag a comment"""
        try:
            comment = get_object_or_404(TicketComment, id=comment_id, ticket_id=pk)
            comment.flagged = not comment.flagged
            comment.save()

            action = "flagged" if comment.flagged else "unflagged"

            # Send notification email for flagged comments
            if comment.flagged:
                from shared.tasks import send_comment_interaction_notification
                send_comment_interaction_notification.delay(
                    ticket_id=pk,
                    comment_id=comment_id,
                    interaction_type='flagged',
                    actor_id=request.user.id
                )

            return Response({
                "message": f"Comment {action} successfully",
                "flagged": comment.flagged
            }, status=status.HTTP_200_OK)

        except TicketComment.DoesNotExist:
            return Response({"message": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='comments/(?P<comment_id>\d+)/like')
    def like_comment(self, request, pk=None, comment_id=None):
        """Like or unlike a comment"""
        try:
            comment = get_object_or_404(TicketComment, id=comment_id, ticket_id=pk)

            # Check if user already liked this comment
            existing_like = CommentLike.objects.filter(comment=comment, user=request.user).first()

            if existing_like:
                # Unlike: remove the like
                existing_like.delete()
                comment.likes_count = max(0, comment.likes_count - 1)
                comment.save()
                return Response({
                    "message": "Comment unliked successfully",
                    "liked": False,
                    "likes_count": comment.likes_count
                }, status=status.HTTP_200_OK)
            else:
                # Like: create new like
                CommentLike.objects.create(comment=comment, user=request.user)
                comment.likes_count += 1
                comment.save()

                # Send notification email to comment author (if not the liker)
                if comment.author and comment.author.id != request.user.id:
                    from shared.tasks import send_comment_interaction_notification
                    send_comment_interaction_notification.delay(
                        ticket_id=pk,
                        comment_id=comment_id,
                        interaction_type='liked',
                        actor_id=request.user.id
                    )

                return Response({
                    "message": "Comment liked successfully",
                    "liked": True,
                    "likes_count": comment.likes_count
                }, status=status.HTTP_200_OK)

        except TicketComment.DoesNotExist:
            return Response({"message": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='comments/(?P<comment_id>\d+)/reply')
    def reply_to_comment(self, request, pk=None, comment_id=None):
        """Create a reply to a comment"""
        try:
            parent_comment = get_object_or_404(TicketComment, id=comment_id, ticket_id=pk)
            content = request.data.get('content')
            is_internal = request.data.get('is_internal', False)

            if not content:
                return Response({
                    "message": "Reply content is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            reply = CommentReply.objects.create(
                parent_comment=parent_comment,
                author=request.user,
                content=content,
                is_internal=is_internal
            )

            # Send notification email for new reply
            from shared.tasks import send_comment_reply_notification
            send_comment_reply_notification.delay(
                ticket_id=pk,
                comment_id=comment_id,
                reply_id=reply.id,
                replier_id=request.user.id
            )

            return Response({
                "message": "Reply added successfully",
                "reply": {
                    "id": reply.id,
                    "content": reply.content,
                    "author": {
                        "id": reply.author.id,
                        "name": reply.author.full_name(),
                        "email": reply.author.email
                    },
                    "created_at": reply.created_at,
                    "is_internal": reply.is_internal,
                    "likes_count": reply.likes_count
                }
            }, status=status.HTTP_201_CREATED)

        except TicketComment.DoesNotExist:
            return Response({"message": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['put'], url_path='comments/(?P<comment_id>\d+)/edit')
    def edit_comment(self, request, pk=None, comment_id=None):
        """Edit a comment"""
        try:
            comment = get_object_or_404(TicketComment, id=comment_id, ticket_id=pk, author=request.user)
            content = request.data.get('content')

            if not content:
                return Response({
                    "message": "Comment content is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            comment.content = content
            comment.updated_at = timezone.now()
            comment.save()

            return Response({
                "message": "Comment updated successfully",
                "comment": {
                    "id": comment.id,
                    "content": comment.content,
                    "updated_at": comment.updated_at
                }
            }, status=status.HTTP_200_OK)

        except TicketComment.DoesNotExist:
            return Response({"message": "Comment not found or you don't have permission to edit it"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['delete'], url_path='comments/(?P<comment_id>\d+)/delete')
    def delete_comment(self, request, pk=None, comment_id=None):
        """Delete a comment"""
        try:
            comment = get_object_or_404(TicketComment, id=comment_id, ticket_id=pk, author=request.user)
            comment.delete()

            return Response({
                "message": "Comment deleted successfully"
            }, status=status.HTTP_204_NO_CONTENT)

        except TicketComment.DoesNotExist:
            return Response({"message": "Comment not found or you don't have permission to delete it"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='comments/(?P<comment_id>\d+)/replies/(?P<reply_id>\d+)/like')
    def like_comment_reply(self, request, pk=None, comment_id=None, reply_id=None):
        """Like or unlike a comment reply"""
        try:
            reply = get_object_or_404(CommentReply, id=reply_id, parent_comment_id=comment_id)

            # Check if user already liked this reply
            existing_like = CommentReplyLike.objects.filter(reply=reply, user=request.user).first()

            if existing_like:
                # Unlike: remove the like
                existing_like.delete()
                reply.likes_count = max(0, reply.likes_count - 1)
                reply.save()
                return Response({
                    "message": "Reply unliked successfully",
                    "liked": False,
                    "likes_count": reply.likes_count
                }, status=status.HTTP_200_OK)
            else:
                # Like: create new like
                CommentReplyLike.objects.create(reply=reply, user=request.user)
                reply.likes_count += 1
                reply.save()
                return Response({
                    "message": "Reply liked successfully",
                    "liked": True,
                    "likes_count": reply.likes_count
                }, status=status.HTTP_200_OK)

        except CommentReply.DoesNotExist:
            return Response({"message": "Reply not found"}, status=status.HTTP_404_NOT_FOUND)

    def export_tickets(self, request, *args, **kwargs):
        """Export selected tickets to Excel format"""
        try:
            # Get selected ticket IDs from request
            ticket_ids = request.data.get('ticket_ids', [])
            if not ticket_ids:
                return Response({"error": "No tickets selected for export"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Get tickets from database
            tickets = Ticket.objects.for_business().filter(id__in=ticket_ids)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Tickets Export"
            
            # Add headers
            headers = ['ID', 'Title', 'Status', 'Assignee', 'Created Date']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add ticket data
            for row, ticket in enumerate(tickets, 2):
                ws.cell(row=row, column=1, value=ticket.ticket_id)
                ws.cell(row=row, column=2, value=ticket.title)
                ws.cell(row=row, column=3, value=ticket.status)
                ws.cell(row=row, column=4, value=ticket.assigned_to.get_full_name() if ticket.assigned_to else 'Unassigned')
                ws.cell(row=row, column=5, value=ticket.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="tickets_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Save workbook to response
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def bulk_archive(self, request, *args, **kwargs):
        """Archive selected tickets"""
        try:
            ticket_ids = request.data.get('ticket_ids', [])
            if not ticket_ids:
                return Response({"error": "No tickets selected for archiving"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tickets to archived
            updated_count = Ticket.objects.for_business().filter(id__in=ticket_ids).update(is_archived=True)
            
            return Response({
                "message": f"Successfully archived {updated_count} tickets",
                "archived_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def bulk_delete(self, request, *args, **kwargs):
        """Soft delete selected tickets"""
        try:
            ticket_ids = request.data.get('ticket_ids', [])
            if not ticket_ids:
                return Response({"error": "No tickets selected for deletion"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tickets to deleted
            updated_count = Ticket.objects.for_business().filter(id__in=ticket_ids).update(is_deleted=True)
            
            return Response({
                "message": f"Successfully deleted {updated_count} tickets",
                "deleted_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def bulk_unarchive(self, request, *args, **kwargs):
        """Unarchive selected tickets"""
        try:
            ticket_ids = request.data.get('ticket_ids', [])
            if not ticket_ids:
                return Response({"error": "No tickets selected for unarchiving"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tickets to unarchived
            updated_count = Ticket.objects.for_business().filter(id__in=ticket_ids).update(is_archived=False)
            
            return Response({
                "message": f"Successfully unarchived {updated_count} tickets",
                "unarchived_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def bulk_restore(self, request, *args, **kwargs):
        """Restore selected tickets from trash"""
        try:
            ticket_ids = request.data.get('ticket_ids', [])
            if not ticket_ids:
                return Response({"error": "No tickets selected for restoration"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tickets to restored (un-delete)
            updated_count = Ticket.objects.for_business().filter(id__in=ticket_ids).update(is_deleted=False)
            
            return Response({
                "message": f"Successfully restored {updated_count} tickets",
                "restored_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @transaction.atomic
    @action(detail=True, methods=['post'])
    def merge(self, request, *args, **kwargs):
        """
        Merge one or more tickets into this ticket (primary).
        Soft-link strategy: mark source tickets as merged/closed without moving related records.
        """
        primary_id = kwargs.get('pk')
        serializer = TicketMergeRequestSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        source_ids = serializer.validated_data.get('source_ids', [])
        note = serializer.validated_data.get('note', '')

        try:
            primary = Ticket.objects.select_for_update().get(id=primary_id)
        except Ticket.DoesNotExist:
            return Response({"message": "Primary ticket not found"}, status=status.HTTP_404_NOT_FOUND)

        # Prevent self-merge
        if primary.id in source_ids:
            return Response({"message": "Cannot merge a ticket into itself"}, status=status.HTTP_400_BAD_REQUEST)

        # Fetch source tickets with lock
        sources = list(
            Ticket.objects.select_for_update()
            .filter(id__in=source_ids)
        )

        if len(sources) != len(set(source_ids)):
            return Response({"message": "One or more source tickets were not found"}, status=status.HTTP_404_NOT_FOUND)

        # Department visibility restriction for agents
        user = request.user
        if user.role and user.role.name == 'agent':
            agent_departments = user.department.all()
            # primary must be in agent departments
            if primary.department not in agent_departments:
                return Response({"message": "You do not have permission to merge into this ticket"}, status=status.HTTP_403_FORBIDDEN)
            for src in sources:
                if src.department not in agent_departments:
                    return Response({"message": f"No permission to merge ticket {src.ticket_id}"}, status=status.HTTP_403_FORBIDDEN)

        merged_ticket_ids = []

        for src in sources:
            old_status = src.status
            if src.is_merged:
                return Response({"message": f"Ticket {src.ticket_id} is already merged"}, status=status.HTTP_400_BAD_REQUEST)
            # Mark as merged/closed
            src.is_merged = True
            src.merged_into = primary
            src.merged_by = request.user if request.user.is_authenticated else None
            src.merged_at = timezone.now()
            src.merge_note = note
            src.status = 'closed'
            src.resolved_at = timezone.now()
            src.is_sla_paused = True
            src.save()

            # Activity on source
            from tenant.models import TicketActivity
            # First, record a merge event with explicit target ticket info
            TicketActivity.objects.create(
                ticket=src,
                user=request.user,
                activity_type='merged',
                description=f"Merged into {primary.ticket_id}",
                old_value='',
                new_value=primary.ticket_id,
            )
            # Then, record the closure
            TicketActivity.objects.create(
                ticket=src,
                user=request.user,
                activity_type='closed',
                description=f"Ticket merged into {primary.ticket_id}",
                old_value=old_status,
                new_value='closed'
            )

            merged_ticket_ids.append(src.ticket_id)

        # Activity on primary
        if merged_ticket_ids:
            from tenant.models import TicketActivity, TicketComment
            merged_list = ", ".join(merged_ticket_ids)
            TicketActivity.objects.create(
                ticket=primary,
                user=request.user,
                activity_type='merged',
                description=f"Merged tickets: {merged_list}",
                old_value='',
                new_value=merged_list
            )
            TicketComment.objects.create(
                ticket=primary,
                author=request.user if request.user.is_authenticated else None,
                content=f"**Merged tickets**: {merged_list}\n\n{note}" if note else f"**Merged tickets**: {merged_list}",
                is_internal=True
            )

        def _collect_recipients(ticket_obj):
            recipients = []
            if ticket_obj.assigned_to:
                recipients.append(ticket_obj.assigned_to)
            if ticket_obj.created_by:
                recipients.append(ticket_obj.created_by)
            for watcher in ticket_obj.watchers.all():
                if watcher.watcher:
                    recipients.append(watcher.watcher)
            if ticket_obj.department:
                dept_members = ticket_obj.department.get_members()
                recipients.extend(list(dept_members))
            return recipients

        def _dedupe(users):
            seen = set()
            unique = []
            for u in users:
                if not u or not u.id:
                    continue
                if u.id not in seen:
                    seen.add(u.id)
                    unique.append(u)
            return unique

        primary_recipients = []
        source_recipients = []

        # In-app notifications (deduped)
        try:
            merged_list_str = ", ".join(merged_ticket_ids)

            # Notify primary side
            primary_recipients = _dedupe([u for u in _collect_recipients(primary) if u != request.user])

            # Notify source side (aggregate)
            aggregate_source_recipients = []
            for src in sources:
                aggregate_source_recipients.extend(_collect_recipients(src))
            source_recipients = _dedupe([u for u in aggregate_source_recipients if u != request.user])

            if primary_recipients:
                NotificationSettingsService.notify_many(
                    users=primary_recipients,
                    ticket=primary,
                    message=f"Tickets merged into #{primary.ticket_id}: {merged_list_str}",
                    notification_type="ticket_status_changed",
                    metadata={
                        "merged_tickets": merged_ticket_ids,
                        "activity_type": "merged",
                        "note": note,
                    }
                )

            for src in sources:
                src_recipients = _dedupe([
                    u for u in _collect_recipients(src)
                    if u != request.user
                ])
                if not src_recipients:
                    continue
                NotificationSettingsService.notify_many(
                    users=src_recipients,
                    ticket=src,
                    message=f"Ticket #{src.ticket_id} was merged into #{primary.ticket_id}",
                    notification_type="ticket_status_changed",
                    metadata={
                        "merged_into": primary.ticket_id,
                        "activity_type": "merged",
                        "note": note,
                    }
                )
        except Exception as notif_error:
            logger.error(f"Merge notifications failed: {notif_error}")

        # DISABLED: Activity email notifications are turned off
        # try:
        #     # Primary recipients: use NEW_ACTIVITY_NOTICE
        #     email_recipients = set()
        #     for user_obj in primary_recipients:
        #         if getattr(user_obj, "email", None):
        #             email_recipients.add(user_obj.email)
        #     for email in email_recipients:
        #         _send_ticket_notification(primary.id, "NEW_ACTIVITY_NOTICE", email)
        #
        #     for src in sources:
        #         src_email_recipients = set()
        #         for user_obj in _collect_recipients(src):
        #             if getattr(user_obj, "email", None) and user_obj != request.user:
        #                 src_email_recipients.add(user_obj.email)
        #         for email in src_email_recipients:
        #             _send_ticket_notification(src.id, "NEW_ACTIVITY_NOTICE", email)
        # except Exception as email_error:
        #     logger.error(f"Merge email notifications failed: {email_error}")

        return Response({
            "message": "Tickets merged successfully",
            "primary_ticket": primary.ticket_id,
            "merged_tickets": merged_ticket_ids
        }, status=status.HTTP_200_OK)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def get_config(self, request):
        """
        Get ticket configuration for the current business.
        Admin-only endpoint.
        """
        user_role = request.user.role.name if request.user.role else None
        if user_role not in ['admin', 'super_admin']:
            return Response(
                {'error': 'Permission denied. Only admins can access settings.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            config = TicketConfig.objects.for_business().first()
            
            if not config:
                # Return default configuration
                return Response({
                    "id_format": "ITK-{YYYY}-{####}",
                }, status=status.HTTP_200_OK)
            
            return Response({
                "id_format": config.id_format,
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching ticket config: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post', 'put'], permission_classes=[IsAuthenticated])
    def update_config(self, request):
        """
        Create or update ticket configuration for the current business.
        Admin-only endpoint.
        """
        user_role = request.user.role.name if request.user.role else None
        if user_role not in ['admin', 'super_admin']:
            return Response(
                {'error': 'Permission denied. Only admins can modify settings.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            config, created = TicketConfig.objects.update_or_create(
                defaults={
                    "id_format": request.data.get("id_format", "ITK-{YYYY}-{####}"),
                    "updated_by": request.user,
                }
            )
            
            return Response({
                "message": "Ticket settings saved successfully",
                "created": created
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error updating ticket config: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
