# tasks/views.py
import os
import uuid
import re
from datetime import datetime
from io import BytesIO
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db import models
from django.db.models import Q
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from openpyxl import Workbook
import logging

from RNSafarideskBack.settings import FILE_URL
from RNSafarideskBack import settings
from tenant.models.DepartmentModel import Department
from tenant.models.TaskModel import Task, TaskAttachment, TaskComment, TaskReplayAttachment, TaskCommentLike, TaskCommentReply, TaskActivity
from tenant.models.TicketModel import Ticket, TicketActivity
from tenant.models.ConfigModel import TaskConfig
from tenant.serializers.TaskSerializer import TaskAddComment, TaskAssignSerializer, TaskAttachToTicketSerializer, TaskCreateSerializer, TaskDetailSerializer, TaskListSerializer, TaskSerializer, TaskStatusUpdateSerializer
from users.models import Users
from util.Helper import Helper
from django.db import transaction
from shared.services.notification_preferences import NotificationSettingsService
from tenant.views.TicketView import _build_storage_paths

logger = logging.getLogger(__name__)


class TaskViewSet(viewsets.ModelViewSet):
    queryset = Task.objects.all().order_by('-created_at')
    serializer_class = TaskSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        queryset = Task.objects.for_business()
        return queryset

    def get_serializer_class(self):
        if self.action == 'create_task':
            return TaskCreateSerializer
        elif self.action in ['list', 'my_tasks']:
            return TaskListSerializer
        elif self.action == 'assign':
            return TaskAssignSerializer
        elif self.action == 'update_status':
            return TaskStatusUpdateSerializer
        elif self.action == 'attach_to_ticket':
            return TaskAttachToTicketSerializer
        elif self.action == 'retrieve':
            return TaskDetailSerializer
        elif self.action == 'add_comment':
            return TaskAddComment
        return TaskSerializer
    
    def retrieve(self, request, *args, **kwargs):
        try:
            task = Task.objects.get(task_trackid=kwargs.get('track_id'))
        except Task.DoesNotExist:
            return Response({"detail": "Task not found"}, status=status.HTTP_404_NOT_FOUND)

        serializer = self.get_serializer(task)
        return Response(serializer.data)
    


    @transaction.atomic
    def create_task(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        if not serializer.is_valid():
            return Response({
                "message": "Validation failed",
                "errors": serializer.errors
            }, status=status.HTTP_400_BAD_REQUEST)
        
        department_id = serializer.validated_data.get('department_id')
        assigned_to_id = serializer.validated_data.get('assigned_to')
        ticket_id = serializer.validated_data.get('ticket_id')
        
        # Department is required - validate it exists
        try:
            department = Department.objects.get(
                id=department_id
            )
        except Department.DoesNotExist:
            return Response({
                "message": f"No department found with id {department_id}"
            }, status=status.HTTP_404_NOT_FOUND)
        
        # Assigned agent optional - validate only if provided
        agent = None
        if assigned_to_id not in [None, '']:
            try:
                agent = Users.objects.get(
                    id=assigned_to_id
                )
            except Users.DoesNotExist:
                return Response({
                    "message": f"No agent found with id {assigned_to_id}"
                }, status=status.HTTP_404_NOT_FOUND)

        # Ticket link optional - validate only if provided
        ticket = None
        if ticket_id not in [None, '']:
            try:
                ticket = Ticket.objects.for_business().get(id=ticket_id)
            except Ticket.DoesNotExist:
                return Response({
                    "message": f"No ticket found with id {ticket_id}"
                }, status=status.HTTP_404_NOT_FOUND)
        
        # Priority: use provided value or fallback
        priority = serializer.validated_data.get('priority') or 'medium'
        
        # Create task with config-based ID
        task = Task.objects.create(
            title=serializer.validated_data.get('title'),
            description=serializer.validated_data.get('description', ''),
            due_date=serializer.validated_data.get('due_date'),
            assigned_to=agent,
            task_trackid=Helper().generate_task_code(),
            department=department,
            priority=priority,
            linked_ticket=ticket,
        )
        if request.FILES:
            uploaded_files = request.FILES
            
            for uploaded_file_name, uploaded_file in uploaded_files.items():
                try:
                    file_extension = os.path.splitext(uploaded_file.name)[1].lower()
                    
                    # Validate file extension (add your allowed extensions)
                    allowed_extensions = ['.jpg', '.jpeg', '.png', '.pdf', '.doc', '.docx', '.txt']
                    if file_extension not in allowed_extensions:
                        return Response({
                            "error": f"File type {file_extension} not allowed"
                        }, status=status.HTTP_400_BAD_REQUEST)
                    
                    # Generate unique filename
                    unique_filename = f"{uuid.uuid4()}{file_extension}"
                    
                    properties_dir, url_base = _build_storage_paths(None, subfolder='files')

                    # Full file path
                    file_path = os.path.join(properties_dir, unique_filename)
                    
                    # Save file
                    with open(file_path, 'wb+') as destination:
                        for chunk in uploaded_file.chunks():
                            destination.write(chunk)
                    
                    # Generate URL
                    file_url = f"{url_base}/{unique_filename}" if url_base else unique_filename
                    
                    # Create attachment
                    TaskAttachment.objects.create(
                        task=task,
                        file_url=file_url,
                        filename=uploaded_file.name,  # Save original filename
                        description=f"File uploaded for task {task.task_trackid}"
                    )
                    
                except Exception as file_error:
                    return Response({
                        "error": f"Failed to save file {uploaded_file_name}",
                        "details": str(file_error)
                    }, status=status.HTTP_400_BAD_REQUEST)

        # Log task creation activity
        TaskActivity.objects.create(
            task=task,
            user=request.user,
            activity_type='created',
            description=f'{request.user.first_name} {request.user.last_name} created the task',
            old_value='',
            new_value=task.task_status
        )

        # If linked to a ticket, log linkage on both task and ticket
        if ticket:
            TaskActivity.objects.create(
                task=task,
                user=request.user,
                activity_type='attached_to_ticket',
                description=f'{request.user.first_name} {request.user.last_name} attached task to ticket #{ticket.ticket_id}',
                old_value='None',
                new_value=ticket.ticket_id
            )

            TicketActivity.objects.create(
                ticket=ticket,
                user=request.user,
                activity_type='task_linked',
                description=f'Task "{task.title}" was linked to this ticket',
                old_value='',
                new_value=task.task_trackid
            )

        assigned_to_data = None
        if task.assigned_to:
            assigned_to_data = {
                'id': task.assigned_to.id,
                'first_name': task.assigned_to.first_name,
                'last_name': task.assigned_to.last_name,
                'email': task.assigned_to.email
            }

        return Response({
            'message': "Task created successfully.",
            'task': {
                'id': task.id,
                'title': task.title,
                'description': task.description,
                'task_trackid': task.task_trackid,
                'task_status': task.task_status,
                'priority': task.priority,
                'due_date': task.due_date,
                'department': {
                    'id': task.department.id,
                    'name': task.department.name
                },
                'assigned_to': assigned_to_data,
                'created_at': task.created_at,
            }
        }, status=status.HTTP_201_CREATED)
    
    def list(self, request, *args, **kwargs):
        queryset = Task.objects.for_business().order_by('-id')

        # Apply view filters first
        view = request.query_params.get('view', 'all_tasks')

        # Exclude archived and deleted tasks from normal views
        if view not in ['archived_tasks', 'trash_tasks']:
            queryset = queryset.filter(is_archived=False, is_deleted=False)

        if view == 'my_tasks':
            queryset = queryset.filter(assigned_to=request.user)
        elif view == 'open_tasks':
            queryset = queryset.filter(task_status='open')
        elif view == 'in_progress_tasks':
            queryset = queryset.filter(task_status='in_progress')
        elif view == 'completed_tasks':
            queryset = queryset.filter(task_status='completed')
        elif view == 'overdue_tasks':
            from django.utils import timezone
            queryset = queryset.filter(
                due_date__lt=timezone.now().date(),
                task_status__in=['open', 'in_progress']
            )
        elif view == 'archived_tasks':
            queryset = queryset.filter(is_archived=True)
        elif view == 'trash_tasks':
            queryset = queryset.filter(is_deleted=True)

        # Apply additional filters
        status_filter = request.query_params.get('status')
        priority_filter = request.query_params.get('priority')
        search_query = request.query_params.get('search')

        if status_filter and status_filter != 'all':
            queryset = queryset.filter(task_status=status_filter)

        if priority_filter and priority_filter != 'all':
            queryset = queryset.filter(priority=priority_filter)

        if search_query:
            queryset = queryset.filter(
                models.Q(title__icontains=search_query) |
                models.Q(task_trackid__icontains=search_query)
            )

        # Additional field filters
        assigned_to_filter = request.query_params.get('assigned_to')
        if assigned_to_filter:
            if assigned_to_filter == 'unassigned':
                queryset = queryset.filter(assigned_to__isnull=True)
            elif assigned_to_filter != 'all':
                queryset = queryset.filter(assigned_to_id=assigned_to_filter)

        department_filter = request.query_params.get('department')
        if department_filter and department_filter != 'all':
            queryset = queryset.filter(department_id=department_filter)

        # Date range filters
        date_from = request.query_params.get('date_from')
        if date_from:
            try:
                from_date = timezone.datetime.fromisoformat(date_from.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__gte=from_date)
            except (ValueError, AttributeError):
                pass

        date_to = request.query_params.get('date_to')
        if date_to:
            try:
                to_date = timezone.datetime.fromisoformat(date_to.replace('Z', '+00:00'))
                queryset = queryset.filter(created_at__lte=to_date)
            except (ValueError, AttributeError):
                pass

        pagination = request.query_params.get('pagination', 'yes').lower()

        if pagination != 'no':
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        # Unpaginated fallback
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

    def my_tasks(self, request, *args, **kwargs):
        queryset = Task.objects.filter(assigned_to=request.user).order_by('-id')
        pagination = request.query_params.get('pagination', 'yes').lower()

        if pagination != 'no':
            page = self.paginate_queryset(queryset)
            if page is not None:
                serializer = self.get_serializer(page, many=True)
                return self.get_paginated_response(serializer.data)

        # Unpaginated fallback
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


    



    @action(detail=True, methods=['post'], url_path='assign')
    def assign(self, request, pk=None):
        try:
            task = Task.objects.for_business().get(id=pk)
        except Task.DoesNotExist:
            return Response({
                'message': 'Task not found.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        user_id = request.data.get('user_id')
        if not user_id:
            return Response({
                'message': 'User ID is required.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            user = Users.objects.get(
                id=user_id,
                
                status='A'
            )
            
            # Store old assignee for activity log
            old_assignee = task.assigned_to
            old_assignee_name = f'{old_assignee.first_name} {old_assignee.last_name}' if old_assignee else 'Unassigned'
            
            task.assigned_to = user
            task.updated_by = request.user
            task.save()
            
            # Log activity
            activity_type = 'reassigned' if old_assignee else 'assigned'
            new_assignee_name = f'{user.first_name} {user.last_name}'
            description = f'{request.user.first_name} {request.user.last_name} {"reassigned" if old_assignee else "assigned"} the task to {new_assignee_name}'
            
            TaskActivity.objects.create(
                task=task,
                user=request.user,
                activity_type=activity_type,
                description=description,
                old_value=old_assignee_name,
                new_value=new_assignee_name
            )
            
            # Create in-app notification for the assignee
            try:
                NotificationSettingsService.create_in_app_notification(
                    user=user,
                    ticket=task.linked_ticket if task.linked_ticket else None,
                    notification_type='task_assigned',
                    message=f'{request.user.full_name() or request.user.email} assigned you to task: {task.title}',
                    metadata={
                        'task_id': task.id,
                        'assigned_by': request.user.id,
                        'assigned_by_name': request.user.full_name() or request.user.email,
                        'old_assignee': old_assignee.id if old_assignee else None
                    }
                )
            except Exception as notif_error:
                logger.error(f"Failed to create task assignment notification: {notif_error}")
            
            return Response({
                'message': 'Task assigned successfully.',
                'assigned_to': {
                    'id': user.id,
                    'name': f'{user.first_name} {user.last_name}',
                    'email': user.email
                }
            })
        except Users.DoesNotExist:
            return Response({
                'message': f'User with id {user_id} not found in your business.'
            }, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='update-status')
    def update_status(self, request, pk=None):
        try:
            task = Task.objects.for_business().get(id=pk)
        except Task.DoesNotExist:
            return Response({
                'message': 'Task not found.'
            }, status=status.HTTP_404_NOT_FOUND)
        
        status_value = request.data.get('status')
        if not status_value:
            return Response({
                'message': 'Status is required.'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Validate status value
        valid_statuses = [choice[0] for choice in Task.STATUS_CHOICES]
        if status_value not in valid_statuses:
            return Response({
                'message': f'Invalid status value. Must be one of: {", ".join(valid_statuses)}'
            }, status=status.HTTP_400_BAD_REQUEST)
        
        # Store old status for activity log
        old_status = task.task_status
        old_status_display = task.get_task_status_display()
        
        task.task_status = status_value
        task.completed_at = timezone.now() if status_value == 'completed' else None
        task.updated_by = request.user
        task.updated_at = timezone.now()
        task.save()
        
        # Log activity
        new_status_display = dict(Task.STATUS_CHOICES).get(status_value, status_value)
        description = f'{request.user.first_name} {request.user.last_name} changed status from {old_status_display} to {new_status_display}'
        
        TaskActivity.objects.create(
            task=task,
            user=request.user,
            activity_type='status_changed',
            description=description,
            old_value=old_status_display,
            new_value=new_status_display
        )
        
        # Create in-app notifications for relevant users
        try:
            recipients = set()
            
            # Notify task creator if different from current user
            if task.created_by and task.created_by.id != request.user.id:
                recipients.add(task.created_by.id)
            
            # Notify assignee if different from current user
            if task.assigned_to and task.assigned_to.id != request.user.id:
                recipients.add(task.assigned_to.id)
            
            # Create notification for each recipient
            for user_id in recipients:
                try:
                    recipient = Users.objects.get(id=user_id)
                    NotificationSettingsService.create_in_app_notification(
                        user=recipient,
                        ticket=task.linked_ticket if task.linked_ticket else None,
                        notification_type='task_status_changed',
                        message=f'{request.user.full_name() or request.user.email} changed task status to {new_status_display}: {task.title}',
                        metadata={
                            'task_id': task.id,
                            'changed_by': request.user.id,
                            'changed_by_name': request.user.full_name() or request.user.email,
                            'old_status': old_status_display,
                            'new_status': new_status_display
                        }
                    )
                except Users.DoesNotExist:
                    logger.warning(f"User with id {user_id} not found for task status notification")
                except Exception as notif_error:
                    logger.error(f"Failed to create task status notification for user {user_id}: {notif_error}")
        except Exception as e:
            logger.error(f"Failed to send task status change notifications: {e}")
        
        return Response({
            'message': 'Status updated successfully.',
            'task': {
                'id': task.id,
                'status': task.task_status,
                'completed_at': task.completed_at
            }
        })

    def partial_update(self, request, pk=None):
        """
        Handle PATCH requests to update task fields like priority, sla, etc.
        """
        try:
            task = Task.objects.for_business().get(id=pk)
        except Task.DoesNotExist:
            return Response({
                'message': 'Task not found.'
            }, status=status.HTTP_404_NOT_FOUND)

        # Track changes for activity log
        changes = []

        # Handle priority update
        if 'priority' in request.data:
            new_priority = request.data.get('priority')
            valid_priorities = ['low', 'medium', 'high', 'critical']
            if new_priority.lower() not in valid_priorities:
                return Response({
                    'message': f'Invalid priority. Must be one of: {", ".join(valid_priorities)}'
                }, status=status.HTTP_400_BAD_REQUEST)
            
            old_priority = task.priority
            task.priority = new_priority.lower()
            changes.append({
                'field': 'priority',
                'old': old_priority,
                'new': new_priority.lower()
            })

        # Handle SLA update
        if 'sla_id' in request.data:
            from tenant.models import SLA
            sla_id = request.data.get('sla_id')
            if sla_id:
                try:
                    sla = SLA.objects.get(id=sla_id)
                    old_sla = task.sla.name if task.sla else 'None'
                    task.sla = sla
                    changes.append({
                        'field': 'sla',
                        'old': old_sla,
                        'new': sla.name
                    })
                except SLA.DoesNotExist:
                    return Response({
                        'message': 'SLA not found.'
                    }, status=status.HTTP_404_NOT_FOUND)
            else:
                # Allow removing SLA
                old_sla = task.sla.name if task.sla else 'None'
                task.sla = None
                changes.append({
                    'field': 'sla',
                    'old': old_sla,
                    'new': 'None'
                })

        # Handle description update
        if 'description' in request.data:
            old_description = task.description[:50] + '...' if len(task.description or '') > 50 else task.description
            task.description = request.data.get('description')
            new_description = task.description[:50] + '...' if len(task.description or '') > 50 else task.description
            changes.append({
                'field': 'description',
                'old': old_description,
                'new': new_description
            })

        # Handle title update
        if 'title' in request.data:
            old_title = task.title
            task.title = request.data.get('title')
            changes.append({
                'field': 'title',
                'old': old_title,
                'new': task.title
            })

        # Handle due date update
        if 'due_date' in request.data:
            old_due_date = str(task.due_date) if task.due_date else 'None'
            task.due_date = request.data.get('due_date')
            new_due_date = str(task.due_date) if task.due_date else 'None'
            changes.append({
                'field': 'due_date',
                'old': old_due_date,
                'new': new_due_date
            })

        if not changes:
            return Response({
                'message': 'No valid fields to update.'
            }, status=status.HTTP_400_BAD_REQUEST)

        task.updated_by = request.user
        task.updated_at = timezone.now()
        task.save()

        # Log activities for each change
        for change in changes:
            description = f'{request.user.first_name} {request.user.last_name} changed {change["field"]} from {change["old"]} to {change["new"]}'
            TaskActivity.objects.create(
                task=task,
                user=request.user,
                activity_type=f'{change["field"]}_changed',
                description=description,
                old_value=str(change['old']),
                new_value=str(change['new'])
            )

        return Response({
            'message': 'Task updated successfully.',
            'task': TaskListSerializer(task).data
        })


    @action(detail=True, methods=['post'], url_path='attach-to-ticket')
    def attach_to_ticket(self, request, pk=None):
        try:
            # Get task and ensure it belongs to the user's business
            task = Task.objects.for_business().get(id=pk)
        except Task.DoesNotExist:
            return Response({'error': 'Task not found.'}, status=status.HTTP_404_NOT_FOUND)

        ticket_id = request.data.get('ticket_id')
        if not ticket_id:
            return Response({'error': 'Ticket ID is required.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            # Get ticket and ensure it belongs to the user's business
            ticket = Ticket.objects.for_business().get(id=ticket_id)
        except Ticket.DoesNotExist:
            return Response({'error': 'Ticket not found.'}, status=status.HTTP_404_NOT_FOUND)

        # Store old ticket for activity log
        old_ticket = task.linked_ticket
        old_ticket_id = old_ticket.ticket_id if old_ticket else None
        
        # Attach task to ticket
        task.linked_ticket = ticket
        task.save()
        
        # Log activity on task
        description = f'{request.user.first_name} {request.user.last_name} attached task to ticket #{ticket.ticket_id}'
        
        TaskActivity.objects.create(
            task=task,
            user=request.user,
            activity_type='attached_to_ticket',
            description=description,
            old_value=old_ticket_id or 'None',
            new_value=ticket.ticket_id
        )
        
        # Also create TicketActivity so task appears in ticket's activity stream
        TicketActivity.objects.create(
            ticket=ticket,
            user=request.user,
            activity_type='task_linked',
            description=f'Task "{task.title}" was linked to this ticket',
            old_value='',
            new_value=task.task_trackid  # Store task track ID for link generation
        )

        return Response({'message': 'Task attached to ticket successfully.'})
    
    @transaction.atomic    
    def add_comment(self, request, *args, **kwargs):
        task_id = kwargs.get('id')

        # if not task_id:
        #     return Response({
        #         "message": "Task ID is required"
        #     }, status=status.HTTP_400_BAD_REQUEST)
        
        try:
            task = get_object_or_404(Task, id=task_id)
            comment = request.data.get('comment', '')

            # Convert string boolean to actual boolean
            is_internal = str(request.data.get("is_internal", "false")).lower() == "true"

            # Add the comment to the ticket
            com = task.comments.create(
                task=task,
                author=request.user,
                content=comment,
                updated_by=request.user,
                is_internal=is_internal
            )

            if request.FILES:
                uploaded_files = request.FILES

                for uploaded_file_name, uploaded_file in uploaded_files.items():
                    try:
                        file_extension = os.path.splitext(uploaded_file.name)[1].lower()

                        # Generate unique filename
                        unique_filename = f"{uuid.uuid4()}{file_extension}"

                        properties_dir, url_base = _build_storage_paths(None, subfolder='files')

                        # Full file path
                        file_path = os.path.join(properties_dir, unique_filename)

                        # Save file
                        with open(file_path, 'wb+') as destination:
                            for chunk in uploaded_file.chunks():
                                destination.write(chunk)

                        # Generate URL (defensive: ensure FILE_URL exists)
                        file_url_base = url_base or FILE_URL or getattr(settings, 'FILE_BASE_URL', '')
                        if file_url_base:
                            file_url = f"{file_url_base.rstrip('/')}/{unique_filename}"
                        else:
                            file_url = unique_filename

                        # Create attachment
                        TaskReplayAttachment.objects.create(
                            comment=com,
                            file_url=file_url,
                            filename=uploaded_file.name  # Save original filename
                        )


                    except Exception as file_error:
                        # Log and return a clear error so production won't 500 silently
                        import logging as _logging
                        _logging.getLogger(__name__).exception("Error saving file %s", uploaded_file_name)
                        return Response({
                            "error": f"Failed to save file {uploaded_file_name}",
                            "details": str(file_error)
                        }, status=status.HTTP_400_BAD_REQUEST)


            # if ticket.assigned_to != request.user:
            #     send_comment_notification.delay(
            #         ticket_id=ticket.id,
            #         comment_id=com.id,
            #         author_email=request.user.email
            #     )
            
            # Notify task creator, assigned user, and watchers
            try:
                recipients = set()
                
                # Add task creator
                if task.created_by:
                    recipients.add(task.created_by.id)
                
                # Add assigned user
                if task.assigned_to:
                    recipients.add(task.assigned_to.id)
                
                # Add watchers (if the Task model has watchers - similar to Ticket)
                # if hasattr(task, 'watchers'):
                #     for watcher in task.watchers.all():
                #         recipients.add(watcher.watcher.id)
                
                # Remove the comment author to avoid self-notification
                recipients.discard(request.user.id)
                
                # Create notifications for all recipients
                for user_id in recipients:
                    try:
                        user = Users.objects.get(id=user_id)
                        NotificationSettingsService.create_in_app_notification(
                            user=user,
                            ticket=task.linked_ticket if task.linked_ticket else None,
                            notification_type='task_comment',
                            message=f'{request.user.full_name() or request.user.email} commented on task: {task.title}',
                            metadata={
                                'comment_id': com.id,
                                'task_id': task.id,
                                'comment_by': request.user.id,
                                'comment_by_name': request.user.full_name() or request.user.email,
                                'is_internal': is_internal
                            }
                        )
                    except Users.DoesNotExist:
                        logger.warning(f"User with id {user_id} not found for notification")
                        continue
                    except Exception as notif_error:
                        logger.error(f"Failed to create notification for user {user_id}: {notif_error}")
                        continue
            except Exception as e:
                logger.error(f"Failed to send task comment notifications: {e}")
            
            # Detect @mentions and create notifications
            try:
                if comment and request.user:
                    # Extract all @mentions from the comment
                    mentions = re.findall(r'@([A-Za-z]+(?:\s+[A-Za-z]+)*)', comment)
                    
                    for mention_name in mentions:
                        mention_name = mention_name.strip()
                        if not mention_name:
                            continue
                        
                        # Try to find user by full name, first name, or last name
                        mentioned_users = Users.objects.filter(
                            Q(full_name__icontains=mention_name) |
                            Q(first_name__icontains=mention_name) |
                            Q(last_name__icontains=mention_name)
                        )
                        
                        # Create notification for each matched user
                        for mentioned_user in mentioned_users:
                            # Don't notify the author about their own mention
                            if mentioned_user.id != request.user.id:
                                try:
                                    NotificationSettingsService.create_in_app_notification(
                                        user=mentioned_user,
                                        ticket=task.linked_ticket if task.linked_ticket else None,
                                        notification_type='task_mention',
                                        message=f'{request.user.full_name() or request.user.email} mentioned you in task: {task.title}',
                                        metadata={
                                            'comment_id': com.id,
                                            'task_id': task.id,
                                            'mentioned_by': request.user.id,
                                            'mentioned_by_name': request.user.full_name() or request.user.email
                                        }
                                    )
                                except Exception as notif_error:
                                    logger.error(f"Failed to create mention notification for user {mentioned_user.id}: {notif_error}")
                                    continue
            except Exception as e:
                logger.error(f"Failed to process @mentions: {e}")


            return Response({
                "message": "Comment added successfully"
            }, status=status.HTTP_201_CREATED)
        
        except Task.DoesNotExist:
            return Response({"message": "Task not found"}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            import logging as _logging
            _logging.getLogger(__name__).exception("Unhandled error in task add_comment: %s", e)
            return Response({
                "message": "An unexpected error occurred while adding the comment",
                "details": str(e)
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def get_activity_stream(self, request, *args, **kwargs):
        task_id = kwargs.get('task_id')

        try:
            # Get task without prefetch to avoid cached queries
            task = Task.objects.select_related(
                'department',
                'assigned_to',
                'linked_ticket'
            ).get(
                id=task_id
                # business=request.user.business
            )
            
            # Query comments directly to ensure fresh data
            comments = TaskComment.objects.filter(
                task=task
            ).select_related('author').prefetch_related(
                'attachment',
                'replies__author'
            ).order_by('created_at')
        except Task.DoesNotExist:
            return Response({
                "error": "Task not found"
            }, status=status.HTTP_404_NOT_FOUND)

        # Collect all activities (comments + system events)
        all_activities = []
        
        # Add comments from fresh query
        for comment in comments:
            attachments = []
            for a in comment.attachment.all():
                attachments.append({
                    "id": a.id,
                    "file_url": a.file_url,
                    "filename": a.filename or a.file_url.split('/')[-1],  # Use filename or extract from URL
                })

            replies_data = []
            for reply in comment.replies.all():
                replies_data.append({
                    "id": reply.id,
                    "content": reply.content,
                    "created_at": reply.created_at,
                    "updated_at": reply.updated_at,
                    "is_internal": reply.is_internal,
                    "likes_count": reply.likes_count,
                    "author": {
                        "id": reply.author.id,
                        "avatar": reply.author.avatar_url if reply.author else "",
                        "name": reply.author.full_name() if reply.author else 'Anonymous User',
                        "email": reply.author.email if reply.author else 'Anonymous User',
                    }
                })

            comment_data = {
                "id": comment.id,
                "type": "comment",
                "content": comment.content,
                "created_at": comment.created_at,
                "updated_at": comment.updated_at,
                "is_internal": comment.is_internal,
                "likes_count": comment.likes_count if hasattr(comment, 'likes_count') else 0,
                "attachments": attachments if attachments else None,
                "author": {
                    "id": comment.author.id,
                    "avatar": comment.author.avatar_url if comment.author else "",
                    "name": comment.author.full_name() if comment.author else 'Anonymous User',
                    "email": comment.author.email if comment.author else 'Anonymous User',
                } if comment.author else {
                    "id": 0,
                    "avatar": "",
                    "name": "Anonymous",
                    "email": "",
                },
                "replies": replies_data,
                "_sort_timestamp": comment.created_at
            }
            all_activities.append(comment_data)
        
        # Query task activities directly for fresh data
        activities = TaskActivity.objects.filter(
            task=task
        ).select_related('user').order_by('-timestamp')
        
        # Add system events (task activities)
        for activity in activities:
            activity_data = {
                "id": activity.id,
                "type": "system_event",
                "activity_type": activity.activity_type,
                "user": {
                    "id": activity.user.id,
                    "avatar": activity.user.avatar_url if activity.user else "",
                    "name": activity.user.full_name() if activity.user else 'System',
                    "email": activity.user.email if activity.user else '',
                },
                "timestamp": activity.timestamp,
                "description": activity.description,
                "old_value": activity.old_value,
                "new_value": activity.new_value,
                "_sort_timestamp": activity.timestamp
            }
            all_activities.append(activity_data)
        
        # Sort by timestamp (chronological order)
        all_activities.sort(key=lambda x: x['_sort_timestamp'])
        
        # Remove the _sort_timestamp helper field
        for activity in all_activities:
            del activity['_sort_timestamp']
        
        return Response(all_activities, status=status.HTTP_200_OK)

    @action(detail=True, methods=['post'], url_path='comments/(?P<comment_id>\d+)/like')
    def like_comment(self, request, pk=None, comment_id=None):
        """Like or unlike a task comment"""
        try:
            comment = get_object_or_404(TaskComment, id=comment_id, task_id=pk)

            # Check if user already liked this comment
            existing_like = TaskCommentLike.objects.filter(comment=comment, user=request.user).first()

            if existing_like:
                # Unlike: remove the like
                existing_like.delete()
                comment.likes_count = max(0, comment.likes_count - 1)
                comment.save()
                return Response({
                    "message": "Comment unliked successfully",
                    "liked": False,
                    "likes_count": comment.likes_count
                }, status=status.HTTP_200_OK)
            else:
                # Like: create new like
                TaskCommentLike.objects.create(comment=comment, user=request.user)
                comment.likes_count += 1
                comment.save()
                return Response({
                    "message": "Comment liked successfully",
                    "liked": True,
                    "likes_count": comment.likes_count
                }, status=status.HTTP_200_OK)

        except TaskComment.DoesNotExist:
            return Response({"message": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='comments/(?P<comment_id>\d+)/flag')
    def flag_comment(self, request, pk=None, comment_id=None):
        """Flag or unflag a task comment"""
        try:
            comment = get_object_or_404(TaskComment, id=comment_id, task_id=pk)

            # Toggle the flagged status
            comment.flagged = not comment.flagged
            comment.save()

            return Response({
                "message": f"Comment {'flagged' if comment.flagged else 'unflagged'} successfully",
                "flagged": comment.flagged
            }, status=status.HTTP_200_OK)

        except TaskComment.DoesNotExist:
            return Response({"message": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='comments/(?P<comment_id>\d+)/reply')
    def reply_to_comment(self, request, pk=None, comment_id=None):
        """Create a reply to a task comment"""
        try:
            parent_comment = get_object_or_404(TaskComment, id=comment_id, task_id=pk)
            content = request.data.get('content')
            is_internal = request.data.get('is_internal', False)

            if not content:
                return Response({
                    "message": "Reply content is required"
                }, status=status.HTTP_400_BAD_REQUEST)

            reply = TaskCommentReply.objects.create(
                parent_comment=parent_comment,
                author=request.user,
                content=content,
                is_internal=is_internal
            )

            return Response({
                "message": "Reply added successfully",
                "reply": {
                    "id": reply.id,
                    "content": reply.content,
                    "author": {
                        "id": reply.author.id,
                        "name": reply.author.full_name(),
                        "email": reply.author.email
                    },
                    "created_at": reply.created_at,
                    "is_internal": reply.is_internal,
                    "likes_count": reply.likes_count
                }
            }, status=status.HTTP_201_CREATED)

        except TaskComment.DoesNotExist:
            return Response({"message": "Comment not found"}, status=status.HTTP_404_NOT_FOUND)

    def export_tasks(self, request, *args, **kwargs):
        """Export selected tasks to Excel format"""
        try:
            # Get selected task IDs from request
            task_ids = request.data.get('task_ids', [])
            if not task_ids:
                return Response({"error": "No tasks selected for export"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Get tasks from database
            tasks = Task.objects.for_business().filter(id__in=task_ids)
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Tasks Export"
            
            # Add headers
            headers = ['ID', 'Title', 'Status', 'Assignee', 'Created Date']
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add task data
            for row, task in enumerate(tasks, 2):
                ws.cell(row=row, column=1, value=task.task_trackid)
                ws.cell(row=row, column=2, value=task.title)
                ws.cell(row=row, column=3, value=task.task_status)
                ws.cell(row=row, column=4, value=task.assigned_to.get_full_name() if task.assigned_to else 'Unassigned')
                ws.cell(row=row, column=5, value=task.created_at.strftime('%Y-%m-%d %H:%M:%S'))
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="tasks_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            # Save workbook to response
            wb.save(response)
            return response
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def bulk_archive(self, request, *args, **kwargs):
        """Archive selected tasks"""
        try:
            task_ids = request.data.get('task_ids', [])
            if not task_ids:
                return Response({"error": "No tasks selected for archiving"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tasks to archived
            updated_count = Task.objects.for_business().filter(id__in=task_ids).update(is_archived=True)
            
            return Response({
                "message": f"Successfully archived {updated_count} tasks",
                "archived_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def bulk_delete(self, request, *args, **kwargs):
        """Soft delete selected tasks"""
        try:
            task_ids = request.data.get('task_ids', [])
            if not task_ids:
                return Response({"error": "No tasks selected for deletion"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tasks to deleted
            updated_count = Task.objects.for_business().filter(id__in=task_ids).update(is_deleted=True)
            
            return Response({
                "message": f"Successfully deleted {updated_count} tasks",
                "deleted_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def bulk_unarchive(self, request, *args, **kwargs):
        """Unarchive selected tasks"""
        try:
            task_ids = request.data.get('task_ids', [])
            if not task_ids:
                return Response({"error": "No tasks selected for unarchiving"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tasks to unarchived
            updated_count = Task.objects.for_business().filter(id__in=task_ids).update(is_archived=False)
            
            return Response({
                "message": f"Successfully unarchived {updated_count} tasks",
                "unarchived_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    def task_counts(self, request, *args, **kwargs):
        """Return counts of tasks for the sidebar views"""
        
        base_queryset = Task.objects.for_business()
        
        # Exclude archived and deleted from active counts
        active_queryset = base_queryset.filter(is_archived=False, is_deleted=False)
        my_active_queryset = active_queryset.filter(assigned_to=request.user)
        
        # For archived tasks (exclude deleted)
        archived_queryset = base_queryset.filter(is_archived=True, is_deleted=False)
        
        # For trash/deleted tasks
        trash_queryset = base_queryset.filter(is_deleted=True)
        
        counts = {
            # All tasks
            'all_tasks': active_queryset.count(),
            
            # My tasks
            'my_tasks': my_active_queryset.count(),
            
            # Task status counts
            'open_tasks': active_queryset.filter(task_status='open').count(),
            'in_progress_tasks': active_queryset.filter(task_status='in_progress').count(),
            'completed_tasks': active_queryset.filter(task_status='completed').count(),
            
            # Overdue tasks (due_date in the past)
            'overdue_tasks': active_queryset.filter(due_date__lt=timezone.now()).count(),
            
            # Archived and Trash
            'archived_tasks': archived_queryset.count(),
            'trash_tasks': trash_queryset.count(),
        }
        
        return Response(counts, status=status.HTTP_200_OK)

    def bulk_restore(self, request, *args, **kwargs):
        """Restore selected tasks from trash"""
        try:
            task_ids = request.data.get('task_ids', [])
            if not task_ids:
                return Response({"error": "No tasks selected for restoration"}, status=status.HTTP_400_BAD_REQUEST)
            
            # Update tasks to restored (un-delete)
            updated_count = Task.objects.for_business().filter(id__in=task_ids).update(is_deleted=False)
            
            return Response({
                "message": f"Successfully restored {updated_count} tasks",
                "restored_count": updated_count
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def get_config(self, request):
        """
        Get task configuration for the current business.
        Admin-only endpoint.
        """
        user_role = request.user.role.name if request.user.role else None
        if user_role not in ['admin', 'super_admin']:
            return Response(
                {'error': 'Permission denied. Only admins can access settings.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            config = TaskConfig.objects.for_business().first()
            
            if not config:
                # Return default configuration
                return Response({
                    "id_format": "TSK-{YYYY}-{####}",
                }, status=status.HTTP_200_OK)
            
            return Response({
                "id_format": config.id_format,
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error fetching task config: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post', 'put'], permission_classes=[IsAuthenticated])
    def update_config(self, request):
        """
        Create or update task configuration for the current business.
        Admin-only endpoint.
        """
        user_role = request.user.role.name if request.user.role else None
        if user_role not in ['admin', 'super_admin']:
            return Response(
                {'error': 'Permission denied. Only admins can modify settings.'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        try:
            config, created = TaskConfig.objects.update_or_create(
                
                defaults={
                    "id_format": request.data.get("id_format", "TSK-{YYYY}-{####}"),
                    "updated_by": request.user,
                }
            )
            
            return Response({
                "message": "Task settings saved successfully",
                "created": created
            }, status=status.HTTP_200_OK)
            
        except Exception as e:
            logger.error(f"Error updating task config: {e}")
            return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
